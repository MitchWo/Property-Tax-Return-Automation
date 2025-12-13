"""Integration tests for Phase 3 Transaction Processing."""
import asyncio
import pytest
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker

from app.config import settings
from app.database import Base
from app.models.db_models import (
    Client, TaxReturn, Document, Transaction, TransactionSummary,
    TransactionPattern, CategoryFeedback, TaxRule, PLRowMapping,
    PropertyType, TaxReturnStatus, DocumentStatus
)


# =============================================================================
# FIXTURES
# =============================================================================

@pytest.fixture(scope="session")
def event_loop():
    """Create event loop for async tests."""
    loop = asyncio.get_event_loop_policy().new_event_loop()
    yield loop
    loop.close()


@pytest.fixture(scope="session")
async def engine():
    """Create test database engine."""
    # Use the same database but in a transaction we'll rollback
    engine = create_async_engine(settings.DATABASE_URL, echo=False)
    yield engine
    await engine.dispose()


@pytest.fixture
async def db_session(engine):
    """Create a test database session with rollback."""
    async_session = async_sessionmaker(
        engine, class_=AsyncSession, expire_on_commit=False
    )

    async with async_session() as session:
        yield session
        await session.rollback()


@pytest.fixture
async def sample_client(db_session: AsyncSession):
    """Create a sample client."""
    client = Client(name="Test Client Ltd")
    db_session.add(client)
    await db_session.commit()
    await db_session.refresh(client)
    return client


@pytest.fixture
async def sample_tax_return(db_session: AsyncSession, sample_client: Client):
    """Create a sample tax return."""
    tax_return = TaxReturn(
        client_id=sample_client.id,
        property_address="123 Test Street, Auckland 1010",
        tax_year="FY25",
        property_type=PropertyType.EXISTING,
        gst_registered=False,
        year_of_ownership=2,
        status=TaxReturnStatus.PENDING
    )
    db_session.add(tax_return)
    await db_session.commit()
    await db_session.refresh(tax_return)
    return tax_return


@pytest.fixture
async def sample_document(db_session: AsyncSession, sample_tax_return: TaxReturn):
    """Create a sample document."""
    document = Document(
        tax_return_id=sample_tax_return.id,
        original_filename="test_bank_statement.csv",
        stored_filename=f"{uuid4()}_test_bank_statement.csv",
        file_path="/tmp/test_bank_statement.csv",
        mime_type="text/csv",
        file_size=1024,
        document_type="bank_statement",
        status=DocumentStatus.CLASSIFIED
    )
    db_session.add(document)
    await db_session.commit()
    await db_session.refresh(document)
    return document


# =============================================================================
# DATABASE MODEL TESTS
# =============================================================================

class TestDatabaseModels:
    """Test database models exist and work correctly."""

    @pytest.mark.asyncio
    async def test_client_creation(self, db_session: AsyncSession):
        """Test client model creation."""
        client = Client(name="New Test Client")
        db_session.add(client)
        await db_session.commit()

        assert client.id is not None
        assert client.name == "New Test Client"
        assert client.created_at is not None

    @pytest.mark.asyncio
    async def test_tax_return_creation(self, db_session: AsyncSession, sample_client: Client):
        """Test tax return model creation."""
        tax_return = TaxReturn(
            client_id=sample_client.id,
            property_address="456 Test Ave",
            tax_year="FY25",
            property_type=PropertyType.NEW_BUILD,
            gst_registered=True,
            year_of_ownership=1,
            status=TaxReturnStatus.PENDING
        )
        db_session.add(tax_return)
        await db_session.commit()

        assert tax_return.id is not None
        assert tax_return.property_type == PropertyType.NEW_BUILD
        assert tax_return.gst_registered is True

    @pytest.mark.asyncio
    async def test_transaction_creation(self, db_session: AsyncSession, sample_tax_return: TaxReturn, sample_document: Document):
        """Test transaction model creation."""
        transaction = Transaction(
            tax_return_id=sample_tax_return.id,
            document_id=sample_document.id,
            transaction_date=date(2024, 6, 15),
            description="LOAN INTEREST CHARGED",
            amount=Decimal("-523.45"),
            category_code="interest",
            is_deductible=True,
            deductible_percentage=Decimal("80.00"),
            confidence=Decimal("0.95"),
            categorization_source="yaml_pattern",
            needs_review=False
        )
        db_session.add(transaction)
        await db_session.commit()

        assert transaction.id is not None
        assert transaction.amount == Decimal("-523.45")
        assert transaction.deductible_percentage == Decimal("80.00")

    @pytest.mark.asyncio
    async def test_transaction_summary_creation(self, db_session: AsyncSession, sample_tax_return: TaxReturn):
        """Test transaction summary model creation."""
        summary = TransactionSummary(
            tax_return_id=sample_tax_return.id,
            category_code="interest",
            transaction_count=24,
            gross_amount=Decimal("12500.00"),
            deductible_amount=Decimal("10000.00"),
            gst_amount=Decimal("0.00")
        )
        db_session.add(summary)
        await db_session.commit()

        assert summary.id is not None
        assert summary.transaction_count == 24

    @pytest.mark.asyncio
    async def test_transaction_pattern_creation(self, db_session: AsyncSession):
        """Test transaction pattern model creation."""
        pattern = TransactionPattern(
            pattern_type="exact_payee",
            payee_name="auckland council",
            category_code="rates",
            confidence=Decimal("0.95"),
            match_count=5,
            created_by="test_user"
        )
        db_session.add(pattern)
        await db_session.commit()

        assert pattern.id is not None
        assert pattern.match_count == 5

    @pytest.mark.asyncio
    async def test_tax_rule_exists(self, db_session: AsyncSession):
        """Test tax rules are seeded."""
        result = await db_session.execute(
            select(TaxRule).where(TaxRule.rule_type == "interest_deductibility")
        )
        rules = result.scalars().all()

        # Should have rules for different years
        assert len(rules) > 0

    @pytest.mark.asyncio
    async def test_pl_mappings_exist(self, db_session: AsyncSession):
        """Test P&L mappings are seeded."""
        result = await db_session.execute(select(PLRowMapping))
        mappings = result.scalars().all()

        # Should have all our standard mappings
        assert len(mappings) >= 30

        # Check specific mapping
        interest_mapping = next((m for m in mappings if m.category_code == "interest"), None)
        assert interest_mapping is not None
        assert interest_mapping.pl_row == 26


# =============================================================================
# YAML RULES TESTS
# =============================================================================

class TestYAMLRules:
    """Test YAML categorization rules loading and matching."""

    def test_load_categorization_rules(self):
        """Test loading categorization rules."""
        from app.rules.loader import load_categorization_rules

        rules = load_categorization_rules()

        assert rules is not None
        assert "payees" in rules
        assert "patterns" in rules
        assert "keywords" in rules

    def test_load_bank_parsers(self):
        """Test loading bank parser configs."""
        from app.rules.loader import load_bank_parsers

        parsers = load_bank_parsers()

        assert parsers is not None
        assert "asb" in parsers
        assert "anz" in parsers
        assert "kiwibank" in parsers

    def test_pattern_matcher_exact_payee(self):
        """Test exact payee matching."""
        from app.rules.loader import PatternMatcher

        matcher = PatternMatcher()

        # Test Auckland Council match
        result = matcher.match("AUCKLAND COUNCIL", other_party="Auckland Council")

        assert result is not None
        assert result["category"] == "rates"
        assert result["confidence"] >= 0.95

    def test_pattern_matcher_regex(self):
        """Test regex pattern matching."""
        from app.rules.loader import PatternMatcher

        matcher = PatternMatcher()

        # Test interest pattern
        result = matcher.match("LOAN INTEREST CHARGED", amount=-500.00)

        assert result is not None
        assert result["category"] == "interest"

    def test_pattern_matcher_no_match(self):
        """Test when no pattern matches."""
        from app.rules.loader import PatternMatcher

        matcher = PatternMatcher()

        result = matcher.match("RANDOM UNKNOWN TRANSACTION")

        # Should return None or low confidence
        assert result is None or result["confidence"] < 0.5


# =============================================================================
# TAX RULES SERVICE TESTS
# =============================================================================

class TestTaxRulesService:
    """Test tax rules service."""

    @pytest.mark.asyncio
    async def test_get_interest_deductibility_existing_fy25(self, db_session: AsyncSession):
        """Test interest deductibility for existing property FY25."""
        from app.services.tax_rules_service import get_tax_rules_service

        service = get_tax_rules_service()
        rate = await service.get_interest_deductibility(db_session, "FY25", "existing")

        assert rate == 80.0  # 80% for existing in FY25

    @pytest.mark.asyncio
    async def test_get_interest_deductibility_new_build(self, db_session: AsyncSession):
        """Test interest deductibility for new build."""
        from app.services.tax_rules_service import get_tax_rules_service

        service = get_tax_rules_service()
        rate = await service.get_interest_deductibility(db_session, "FY25", "new_build")

        assert rate == 100.0  # 100% for new builds

    @pytest.mark.asyncio
    async def test_get_accounting_fee(self, db_session: AsyncSession):
        """Test getting standard accounting fee."""
        from app.services.tax_rules_service import get_tax_rules_service

        service = get_tax_rules_service()
        fee = await service.get_accounting_fee(db_session)

        assert fee == Decimal("862.50")

    @pytest.mark.asyncio
    async def test_get_pl_mapping(self, db_session: AsyncSession):
        """Test getting P&L row mapping."""
        from app.services.tax_rules_service import get_tax_rules_service

        service = get_tax_rules_service()
        mapping = await service.get_pl_row_mapping(db_session, "interest")

        assert mapping is not None
        assert mapping.pl_row == 26
        assert mapping.display_name == "Interest"


# =============================================================================
# TRANSACTION EXTRACTOR TESTS
# =============================================================================

class TestTransactionExtractor:
    """Test transaction extraction from documents."""

    @pytest.mark.asyncio
    async def test_parse_asb_csv(self):
        """Test parsing ASB CSV format."""
        from app.services.transaction_extractor_claude import TransactionExtractorClaude as TransactionExtractor

        # Sample ASB CSV content
        csv_content = """Date,Description,Amount,Balance,Account
2024/06/15,LOAN INTEREST,-523.45,10000.00,123456789
2024/06/15,DEPOSIT RENT,2500.00,12500.00,123456789
2024/06/20,AUCKLAND COUNCIL,-450.00,12050.00,123456789"""

        extractor = TransactionExtractor()

        # Create mock document and tax return
        from unittest.mock import MagicMock
        mock_doc = MagicMock()
        mock_doc.original_filename = "ASB_Statement.csv"

        mock_tax_return = MagicMock()
        mock_tax_return.tax_year = "FY25"

        transactions = extractor._parse_csv(csv_content, "asb")

        assert len(transactions) == 3
        assert transactions[0]["amount"] == -523.45
        assert transactions[1]["amount"] == 2500.00

    @pytest.mark.asyncio
    async def test_parse_amount_formats(self):
        """Test parsing various amount formats."""
        from app.services.transaction_extractor_claude import TransactionExtractorClaude as TransactionExtractor

        extractor = TransactionExtractor()

        assert extractor._parse_amount("$1,234.56") == 1234.56
        assert extractor._parse_amount("-$500.00") == -500.00
        assert extractor._parse_amount("(100.00)") == -100.00
        assert extractor._parse_amount("500.00 DR") == -500.00
        assert extractor._parse_amount("500.00 CR") == 500.00


# =============================================================================
# TRANSACTION CATEGORIZER TESTS
# =============================================================================

class TestTransactionCategorizer:
    """Test transaction categorization."""

    @pytest.mark.asyncio
    async def test_categorize_interest_transaction(self, db_session: AsyncSession, sample_tax_return: TaxReturn):
        """Test categorizing an interest transaction."""
        from app.services.transaction_categorizer import TransactionCategorizer

        categorizer = TransactionCategorizer()

        # Create a mock transaction
        from unittest.mock import MagicMock
        txn = MagicMock()
        txn.description = "LOAN INTEREST CHARGED"
        txn.amount = Decimal("-523.45")
        txn.other_party = None

        result = await categorizer.categorize_transaction(db_session, txn, sample_tax_return)

        assert result["category_code"] == "interest"
        assert result["method"] == "yaml_pattern"
        assert result["confidence"] >= 0.8

    @pytest.mark.asyncio
    async def test_learn_from_feedback(self, db_session: AsyncSession, sample_tax_return: TaxReturn, sample_document: Document):
        """Test learning from user corrections."""
        from app.services.transaction_processor import TransactionProcessor

        processor = TransactionProcessor()

        # Create a transaction
        transaction = Transaction(
            tax_return_id=sample_tax_return.id,
            document_id=sample_document.id,
            transaction_date=date(2024, 6, 15),
            description="QUINOVIC MANAGEMENT FEE",
            other_party="Quinovic",
            amount=Decimal("-175.00"),
            category_code="unknown",
            categorization_method="manual",
            confidence=Decimal("0.50"),
            needs_review=True
        )
        db_session.add(transaction)
        await db_session.commit()
        await db_session.refresh(transaction)

        # Submit correction
        success = await processor.learn_from_feedback(
            db=db_session,
            transaction_id=transaction.id,
            correct_category="agent_fees"
        )

        assert success is True

        # Verify pattern was created
        result = await db_session.execute(
            select(TransactionPattern).where(
                TransactionPattern.payee_name == "Quinovic"
            )
        )
        pattern = result.scalar_one_or_none()

        assert pattern is not None
        assert pattern.category_code == "agent_fees"


# =============================================================================
# WORKBOOK GENERATOR TESTS
# =============================================================================

class TestWorkbookGenerator:
    """Test workbook generation."""

    @pytest.mark.asyncio
    async def test_generate_workbook(self, db_session: AsyncSession, sample_tax_return: TaxReturn, sample_document: Document):
        """Test generating a complete workbook."""
        from app.services.workbook_generator import get_workbook_generator

        generator = get_workbook_generator()

        # Create some transactions
        transactions = [
            Transaction(
                tax_return_id=sample_tax_return.id,
                document_id=sample_document.id,
                transaction_date=date(2024, 6, 15),
                description="LOAN INTEREST",
                amount=Decimal("-500.00"),
                category_code="interest",
                deductible_percentage=Decimal("80"),
                deductible_amount=Decimal("400.00")
            ),
            Transaction(
                tax_return_id=sample_tax_return.id,
                document_id=sample_document.id,
                transaction_date=date(2024, 6, 15),
                description="RENT DEPOSIT",
                amount=Decimal("2500.00"),
                category_code="rental_income"
            ),
        ]

        for txn in transactions:
            db_session.add(txn)

        # Create summaries
        summaries = [
            TransactionSummary(
                tax_return_id=sample_tax_return.id,
                category_code="interest",
                transaction_count=1,
                gross_amount=Decimal("500.00"),
                deductible_amount=Decimal("400.00")
            ),
            TransactionSummary(
                tax_return_id=sample_tax_return.id,
                category_code="rental_income",
                transaction_count=1,
                gross_amount=Decimal("2500.00")
            ),
        ]

        for s in summaries:
            db_session.add(s)

        await db_session.commit()

        # Generate workbook
        filepath = await generator.generate_workbook(db_session, sample_tax_return.id)

        assert filepath.exists()
        assert filepath.suffix == ".xlsx"

        # Verify workbook contents
        from openpyxl import load_workbook
        wb = load_workbook(filepath)

        assert "P&L" in wb.sheetnames
        assert "Rental BS" in wb.sheetnames
        assert "Interest Workings" in wb.sheetnames

        # Clean up
        filepath.unlink()


# =============================================================================
# API ENDPOINT TESTS
# =============================================================================

class TestAPIEndpoints:
    """Test API endpoints."""

    @pytest.mark.asyncio
    async def test_health_check(self):
        """Test health check endpoint."""
        from fastapi.testclient import TestClient
        from app.main import app

        client = TestClient(app)
        response = client.get("/health")

        assert response.status_code == 200
        assert response.json()["status"] == "healthy"

    @pytest.mark.asyncio
    async def test_get_categories(self):
        """Test getting all categories."""
        from fastapi.testclient import TestClient
        from app.main import app

        client = TestClient(app)
        response = client.get("/api/transactions/categories/all")

        assert response.status_code == 200
        categories = response.json()

        assert len(categories) > 0
        assert any(c["category_code"] == "interest" for c in categories)
        assert any(c["category_code"] == "rental_income" for c in categories)


# =============================================================================
# RUN TESTS
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])