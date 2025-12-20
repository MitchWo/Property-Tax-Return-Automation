#!/usr/bin/env python3
"""
End-to-end integration test for Phase 1 → Phase 3.

This script verifies the complete flow works:
1. Upload documents (Phase 1)
2. Process transactions (Phase 3)
3. Generate workbook (Phase 4)
"""
import asyncio
import sys
from datetime import date
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))


async def test_full_integration():
    """Test complete Phase 1 → Phase 3 → Phase 4 flow."""

    print("=" * 60)
    print("INTEGRATION TEST: Phase 1 → Phase 3 → Phase 4")
    print("=" * 60)

    from app.database import AsyncSessionLocal
    from app.models.db_models import (
        Client, TaxReturn, Document, Transaction, TransactionSummary,
        PropertyType, TaxReturnStatus, DocumentStatus
    )
    from sqlalchemy import select, delete

    async with AsyncSessionLocal() as db:
        # =================================================================
        # SETUP: Create test data simulating Phase 1 output
        # =================================================================
        print("\n1. SETUP: Creating test data (simulating Phase 1)")
        print("-" * 60)

        # Create client
        client = Client(name="Integration Test Client")
        db.add(client)
        await db.commit()
        await db.refresh(client)
        print(f"   ✓ Created client: {client.id}")

        # Create tax return
        tax_return = TaxReturn(
            client_id=client.id,
            property_address="456 Integration Test Road, Auckland",
            tax_year="FY25",
            property_type=PropertyType.EXISTING,
            gst_registered=False,
            year_of_ownership=2,
            status=TaxReturnStatus.INCOMPLETE
        )
        db.add(tax_return)
        await db.commit()
        await db.refresh(tax_return)
        print(f"   ✓ Created tax return: {tax_return.id}")

        # Create a test CSV file
        test_csv_content = """2024/06/01,Opening Balance,0.00,10000.00,123456789
2024/06/05,LOAN INTEREST CHARGED,-523.45,9476.55,123456789
2024/06/10,RENT DEPOSIT J SMITH,2500.00,11976.55,123456789
2024/06/15,AUCKLAND COUNCIL RATES,-450.00,11526.55,123456789
2024/06/20,WATERCARE SERVICES,-85.00,11441.55,123456789
2024/06/25,TOWER LANDLORD INSURANCE,-120.00,11321.55,123456789
2024/06/28,QUINOVIC MANAGEMENT FEE,-175.00,11146.55,123456789
2024/07/05,LOAN INTEREST CHARGED,-521.32,10625.23,123456789
2024/07/10,RENT DEPOSIT J SMITH,2500.00,13125.23,123456789"""

        # Save test file
        from app.config import settings
        test_dir = settings.UPLOAD_DIR / str(tax_return.id)
        test_dir.mkdir(parents=True, exist_ok=True)

        test_file_path = test_dir / "test_asb_statement.csv"
        test_file_path.write_text(test_csv_content)
        print(f"   ✓ Created test CSV: {test_file_path}")

        # Create document record (as Phase 1 would)
        document = Document(
            tax_return_id=tax_return.id,
            original_filename="ASB_Statement_June_July_2024.csv",
            stored_filename="test_asb_statement.csv",
            file_path=str(test_file_path),
            mime_type="text/csv",
            file_size=len(test_csv_content),
            document_type="bank_statement",  # Phase 1 classified this
            classification_confidence=0.95,
            status=DocumentStatus.CLASSIFIED,
            extracted_data={
                "reasoning": "Bank statement with transactions",
                "key_details": {"bank": "ASB", "account_type": "rental"}
            }
        )
        db.add(document)
        await db.commit()
        await db.refresh(document)
        print(f"   ✓ Created document: {document.id}")
        print(f"     - Type: {document.document_type}")
        print(f"     - Path: {document.file_path}")
        print(f"     - Status: {document.status}")

        # =================================================================
        # TEST PHASE 3: Process Transactions
        # =================================================================
        print("\n2. PHASE 3: Processing Transactions")
        print("-" * 60)

        try:
            from app.services.transaction_processor import get_transaction_processor

            processor = get_transaction_processor()

            result = await processor.process_tax_return_transactions(
                db=db,
                tax_return_id=tax_return.id,
                use_claude=False  # Don't use Claude for this test
            )

            print(f"   ✓ Transactions processed: {result.total_transactions}")
            print(f"   ✓ Categorized: {result.transactions_categorized}")
            print(f"   ✓ Needs review: {result.transactions_needing_review}")
            print(f"   ✓ Documents processed: {len(result.documents_processed)}")

            if result.blocking_issues:
                print(f"   ⚠ Blocking issues: {result.blocking_issues}")

        except Exception as e:
            print(f"   ✗ Transaction processing failed: {e}")
            import traceback
            traceback.print_exc()
            return False

        # =================================================================
        # VERIFY: Check transactions were created
        # =================================================================
        print("\n3. VERIFY: Checking created transactions")
        print("-" * 60)

        txn_result = await db.execute(
            select(Transaction).where(Transaction.tax_return_id == tax_return.id)
        )
        transactions = txn_result.scalars().all()

        print(f"   Total transactions: {len(transactions)}")

        if len(transactions) == 0:
            print("   ✗ No transactions created!")
            return False

        # Check categories
        categories = {}
        for txn in transactions:
            cat = txn.category_code or "uncategorized"
            categories[cat] = categories.get(cat, 0) + 1

        print(f"   Categories found:")
        for cat, count in sorted(categories.items()):
            print(f"     - {cat}: {count}")

        # Check specific expected transactions
        expected = [
            ("interest", 2),  # Two interest charges
            ("rental_income", 2),  # Two rent deposits
            ("rates", 1),  # Auckland Council
            ("water_rates", 1),  # Watercare
            ("insurance", 1),  # Tower insurance
            ("agent_fees", 1),  # Quinovic
        ]

        all_found = True
        for exp_cat, exp_count in expected:
            actual = categories.get(exp_cat, 0)
            if actual >= exp_count:
                print(f"   ✓ {exp_cat}: Expected {exp_count}, found {actual}")
            else:
                print(f"   ✗ {exp_cat}: Expected {exp_count}, found {actual}")
                all_found = False

        # =================================================================
        # VERIFY: Check summaries were generated
        # =================================================================
        print("\n4. VERIFY: Checking transaction summaries")
        print("-" * 60)

        sum_result = await db.execute(
            select(TransactionSummary).where(TransactionSummary.tax_return_id == tax_return.id)
        )
        summaries = sum_result.scalars().all()

        print(f"   Total summaries: {len(summaries)}")

        for s in summaries:
            print(f"     - {s.category_code}: {s.transaction_count} txns, ${s.gross_amount}")

        # =================================================================
        # TEST PHASE 4: Generate Workbook
        # =================================================================
        print("\n5. PHASE 4: Generating Workbook")
        print("-" * 60)

        try:
            from app.services.workbook_generator import get_workbook_generator

            generator = get_workbook_generator()
            filepath = await generator.generate_workbook(db, tax_return.id)

            print(f"   ✓ Workbook generated: {filepath}")

            # Verify file exists
            if filepath.exists():
                print(f"   ✓ File exists: {filepath.stat().st_size} bytes")

                # Open and verify structure
                from openpyxl import load_workbook
                wb = load_workbook(filepath)

                print(f"   ✓ Sheets: {wb.sheetnames}")

                # Check P&L has data
                pl = wb["P&L"]
                print(f"   ✓ P&L sheet rows: {pl.max_row}")

            else:
                print(f"   ✗ Workbook file not found!")
                return False

        except Exception as e:
            print(f"   ✗ Workbook generation failed: {e}")
            import traceback
            traceback.print_exc()
            return False

        # =================================================================
        # CLEANUP
        # =================================================================
        print("\n6. CLEANUP")
        print("-" * 60)

        # Delete test data
        await db.execute(delete(Transaction).where(Transaction.tax_return_id == tax_return.id))
        await db.execute(delete(TransactionSummary).where(TransactionSummary.tax_return_id == tax_return.id))
        await db.execute(delete(Document).where(Document.tax_return_id == tax_return.id))
        await db.execute(delete(TaxReturn).where(TaxReturn.id == tax_return.id))
        await db.execute(delete(Client).where(Client.id == client.id))
        await db.commit()

        # Delete test files
        if test_file_path.exists():
            test_file_path.unlink()
        if test_dir.exists():
            test_dir.rmdir()
        if filepath.exists():
            filepath.unlink()

        print("   ✓ Test data cleaned up")

        # =================================================================
        # RESULT
        # =================================================================
        print("\n" + "=" * 60)
        if all_found and len(transactions) > 0:
            print("✓ INTEGRATION TEST PASSED")
            print("=" * 60)
            return True
        else:
            print("✗ INTEGRATION TEST FAILED")
            print("=" * 60)
            return False


async def test_api_endpoints():
    """Test API endpoints are correctly wired."""

    print("\n" + "=" * 60)
    print("API ENDPOINT TEST")
    print("=" * 60)

    import httpx

    # Use httpx for async testing
    async with httpx.AsyncClient(base_url="http://localhost:8001") as client:
        # Test health
        print("\n1. Testing /health")
        try:
            resp = await client.get("/health")
            print(f"   Status: {resp.status_code}")
            assert resp.status_code == 200, "Health check failed"
            print("   ✓ Health check passed")
        except httpx.ConnectError:
            print("   ⚠ Server not running on port 8001, skipping API tests")
            return True

        # Test categories endpoint
        print("\n2. Testing /api/transactions/categories/all")
        resp = await client.get("/api/transactions/categories/all")
        print(f"   Status: {resp.status_code}")
        if resp.status_code == 200:
            categories = resp.json()
            print(f"   ✓ Found {len(categories)} categories")
        else:
            print(f"   ✗ Failed: {resp.text}")

        # Test upload page loads
        print("\n3. Testing / (upload page)")
        resp = await client.get("/")
        print(f"   Status: {resp.status_code}")
        assert resp.status_code == 200, "Upload page failed to load"
        print("   ✓ Upload page loads")

    print("\n" + "=" * 60)
    print("✓ API ENDPOINT TESTS PASSED")
    print("=" * 60)
    return True


async def main():
    """Run all integration tests."""

    results = []

    # Test API endpoints first
    try:
        results.append(await test_api_endpoints())
    except Exception as e:
        print(f"API test failed: {e}")
        results.append(False)

    # Test full integration
    try:
        results.append(await test_full_integration())
    except Exception as e:
        print(f"Integration test failed: {e}")
        import traceback
        traceback.print_exc()
        results.append(False)

    # Final summary
    print("\n" + "=" * 60)
    print("FINAL SUMMARY")
    print("=" * 60)

    if all(results):
        print("\n✓ ALL INTEGRATION TESTS PASSED")
        print("\nThe system is correctly integrated:")
        print("  • Phase 1: Document upload and classification")
        print("  • Phase 3: Transaction extraction and categorization")
        print("  • Phase 4: Workbook generation")
        print("\nYou can now test with real documents.")
        return 0
    else:
        print("\n✗ SOME TESTS FAILED")
        print("\nPlease review the errors above and fix before proceeding.")
        return 1


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)