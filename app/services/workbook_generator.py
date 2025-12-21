"""Workbook generator service for creating Lighthouse Financial property tax workbooks.

Sheets generated:
- Profit and Loss (P&L left side, workings right side)
- IRD (compliance checklist)
- FY Summary (pivot by category)
- Bank Statement (transaction detail with category codes)
- Loan Statement (interest/principal breakdown)
- PM Statement (property manager transactions, if applicable)
"""
import logging
import re
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.models.db_models import (
    Document,
    PLRowMapping,
    TaxReturn,
    TaxReturnWorkings,
    Transaction,
    TransactionSummary,
)
from app.services.phase2_ai_brain.workings_models import (
    TaxReturnWorkingsData,
    IncomeWorkings,
    ExpenseWorkings,
    LineItem,
    WorkingsSummary,
)
from app.services.tax_rules_service import get_tax_rules_service

logger = logging.getLogger(__name__)

# Currency format matching Lighthouse template
CURRENCY_FORMAT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
PERCENTAGE_FORMAT = '0%'
DATE_FORMAT = 'mmm-yy'


def create_styles(wb: Workbook):
    """Create named styles for the workbook."""

    # Currency style
    currency_style = NamedStyle(name="currency")
    currency_style.number_format = CURRENCY_FORMAT
    currency_style.alignment = Alignment(horizontal="right")
    wb.add_named_style(currency_style)

    # Percentage style
    percent_style = NamedStyle(name="percent")
    percent_style.number_format = PERCENTAGE_FORMAT
    percent_style.alignment = Alignment(horizontal="right")
    wb.add_named_style(percent_style)

    # Date style (mmm-yy)
    date_style = NamedStyle(name="month_date")
    date_style.number_format = DATE_FORMAT
    date_style.alignment = Alignment(horizontal="left")
    wb.add_named_style(date_style)

    # Bold style
    bold_style = NamedStyle(name="bold")
    bold_style.font = Font(bold=True)
    wb.add_named_style(bold_style)

    # Bold centered
    bold_center_style = NamedStyle(name="bold_center")
    bold_center_style.font = Font(bold=True)
    bold_center_style.alignment = Alignment(horizontal="center")
    wb.add_named_style(bold_center_style)

    # Centered
    center_style = NamedStyle(name="center")
    center_style.alignment = Alignment(horizontal="center")
    wb.add_named_style(center_style)


class WorkbookGenerator:
    """Generate Lighthouse Financial template workbooks from processed transactions."""

    def __init__(self):
        """Initialize workbook generator."""
        self.tax_rules_service = get_tax_rules_service()
        self.output_dir = settings.UPLOAD_DIR / "workbooks"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    async def generate_workbook(
        self,
        db: AsyncSession,
        tax_return_id: UUID
    ) -> Path:
        """
        Generate complete Lighthouse Financial workbook for a tax return.

        Uses AI Brain workings data (TaxReturnWorkings) as the primary data source,
        falling back to TransactionSummary if workings not available.

        Args:
            db: Database session
            tax_return_id: Tax return to generate workbook for

        Returns:
            Path to generated workbook file
        """
        # Load all data
        tax_return = await self._load_tax_return(db, tax_return_id)
        transactions = await self._load_transactions_with_documents(db, tax_return_id)
        summaries = await self._load_summaries(db, tax_return_id)
        pl_mappings = await self._load_pl_mappings(db)

        # Try to load AI Brain workings (preferred data source)
        workings = await self._load_workings(db, tax_return_id)

        # Get tax rules
        interest_deductibility = await self.tax_rules_service.get_interest_deductibility(
            db, tax_return.tax_year, tax_return.property_type.value
        )
        deductibility_rate = interest_deductibility / 100

        # Create workbook
        wb = Workbook()
        create_styles(wb)

        # Remove default sheet
        default_sheet = wb.active

        # Get property short name for sheet naming
        property_short = self._get_property_short_name(tax_return.property_address)

        # Create core sheets
        pl_sheet = wb.create_sheet("Profit and Loss", 0)
        ird_sheet = wb.create_sheet("IRD", 1)
        fy_sheet = wb.create_sheet(f"FY{tax_return.tax_year[-2:]}", 2)

        # Create transaction detail sheets based on available data
        sheet_index = 3

        # Bank statement sheet - show all bank transactions with category codes
        bank_txns = [t for t in transactions if self._is_bank_transaction(t)]
        bank_sheet = None
        if bank_txns:
            bank_sheet = wb.create_sheet(f"{property_short} - Bank", sheet_index)
            sheet_index += 1

        # Loan statement sheet - show interest/principal breakdown
        loan_txns = [t for t in transactions if self._is_loan_transaction(t)]
        loan_sheet = None
        if loan_txns:
            loan_sheet = wb.create_sheet(f"{property_short} - Loans", sheet_index)
            sheet_index += 1

        # PM statement sheet - show property manager transactions
        pm_txns = [t for t in transactions if self._is_pm_transaction(t)]
        pm_sheet = None
        if pm_txns:
            pm_sheet = wb.create_sheet(f"{property_short} - PM", sheet_index)
            sheet_index += 1

        # Remove default sheet
        if default_sheet and default_sheet.title == "Sheet":
            wb.remove(default_sheet)

        # Build context
        context = {
            "tax_return": tax_return,
            "transactions": transactions,
            "summaries": summaries,
            "pl_mappings": pl_mappings,
            "workings": workings,  # AI Brain workings (may be None)
            "interest_deductibility": interest_deductibility,
            "deductibility_rate": deductibility_rate,
        }

        # Build P&L sheet - use workings if available, otherwise fall back to summaries
        if workings:
            logger.info("Using AI Brain workings for workbook generation")
            self._build_profit_loss_sheet_from_workings(pl_sheet, context)
        else:
            logger.warning("No AI Brain workings available, falling back to TransactionSummary")
            self._build_profit_loss_sheet(pl_sheet, context)

        # Build other sheets
        self._build_ird_sheet(ird_sheet, context)
        self._build_fy_summary_sheet(fy_sheet, context)

        # Build transaction detail sheets
        if bank_sheet:
            self._build_bank_statement_sheet(bank_sheet, transactions, tax_return.property_address, tax_return.tax_year)

        if loan_sheet:
            self._build_loan_statement_sheet(loan_sheet, transactions, tax_return.property_address, tax_return.tax_year)

        if pm_sheet:
            self._build_pm_statement_sheet(pm_sheet, transactions, tax_return.property_address, tax_return.tax_year)

        # Set P&L as active
        wb.active = pl_sheet

        # Generate filename
        client_name = self._sanitize_filename(tax_return.client.name)
        year = tax_return.tax_year[-2:]
        filename = f"PTR01_-_Rental_Property_Workbook_-_{client_name}_-_{year}.xlsx"
        filepath = self.output_dir / filename

        # Save
        wb.save(filepath)
        logger.info(f"Generated workbook: {filepath}")

        return filepath

    def _sanitize_filename(self, name: str) -> str:
        """Sanitize string for use in filename."""
        return "".join(c if c.isalnum() or c in "_- " else "_" for c in name).replace(" ", "_")

    def _get_summary(self, summary_by_category: Dict[str, TransactionSummary],
                     primary_code: str, alt_codes: List[str] = None) -> Optional[TransactionSummary]:
        """
        Get summary by primary code, falling back to alternatives.

        Args:
            summary_by_category: Dict mapping category_code to TransactionSummary
            primary_code: Primary category code to look up
            alt_codes: Alternative codes to try if primary not found

        Returns:
            TransactionSummary or None
        """
        summary = summary_by_category.get(primary_code)
        if not summary and alt_codes:
            for alt in alt_codes:
                summary = summary_by_category.get(alt)
                if summary:
                    logger.debug(f"Found {primary_code} via alternative code: {alt}")
                    break
        return summary

    def _check_for_duplicates(self, summaries: List[TransactionSummary]):
        """Detect potential double-counting issues."""
        amounts = {}
        for s in summaries:
            # Use deductible_amount if available, otherwise gross
            amt = float(abs(s.deductible_amount or s.gross_amount or 0))
            if amt > 100:  # Only check significant amounts
                if amt in amounts:
                    logger.warning(f"DUPLICATE AMOUNT ${amt:,.2f}: {amounts[amt]} and {s.category_code}")
                else:
                    amounts[amt] = s.category_code

    def _get_source_code(self, summary: Optional[TransactionSummary], category_code: str = None) -> str:
        """
        Determine source code based on where the data came from.

        Returns: BS (Bank Statement), PM (Property Manager), INV (Invoice),
                 AF (Accounting Fee), SS (Settlement Statement), etc.
        """
        # If summary has source tracking
        if summary and hasattr(summary, 'primary_source') and summary.primary_source:
            source = summary.primary_source.lower()
            if 'bank' in source:
                return 'BS'
            elif 'property_manager' in source or 'pm_statement' in source:
                return 'PM'
            elif 'invoice' in source:
                return 'INV'
            elif 'settlement' in source:
                return 'SS'

        # Use category code from summary or passed parameter
        cat = category_code or (summary.category_code if summary else '')
        cat = (cat or '').lower()

        # Categories that typically come from specific sources
        source_map = {
            'rental_income': 'PM',
            'agent_fees': 'PM',
            'property_management_fees': 'PM',
            'property_management': 'PM',
            'letting_fee': 'PM',
            'mortgage_interest': 'BS',
            'interest_expense': 'BS',
            'interest': 'BS',
            'rates': 'BS',
            'council_rates': 'BS',
            'water_rates': 'BS',
            'body_corporate': 'BS',
            'insurance': 'INV',
            'landlord_insurance': 'INV',
            'accounting_fees': 'AF',
            'consulting_accounting': 'AF',
            'repairs_maintenance': 'PM',
            'repairs': 'PM',
            'maintenance': 'PM',
            'bank_fees': 'BS',
            'depreciation': 'QS',  # Quantity Surveyor
        }

        return source_map.get(cat, 'BS')  # Default to Bank Statement

    def _verify_totals(self, context: Dict[str, Any]) -> Dict[str, Any]:
        """
        Verify calculated totals match expected values.
        Uses same logic as API endpoint /api/transactions/totals/{tax_return_id}
        Returns dict of totals for logging.
        """
        summaries = context["summaries"]
        summary_by_cat = {s.category_code: s for s in summaries}

        def get_income_amount(cats: List[str]) -> Decimal:
            """Get income total (uses gross_amount)."""
            total = Decimal(0)
            for cat in cats:
                s = summary_by_cat.get(cat)
                if s and s.gross_amount and s.gross_amount > 0:
                    total += s.gross_amount
            return total

        def get_expense_amount(cats: List[str]) -> Decimal:
            """Get expense total (uses deductible_amount, fallback to gross)."""
            total = Decimal(0)
            for cat in cats:
                s = summary_by_cat.get(cat)
                if s:
                    # Use deductible_amount if available, otherwise gross_amount
                    amt = s.deductible_amount if s.deductible_amount else s.gross_amount
                    if amt:
                        total += abs(amt)
            return total

        # Calculate expected totals matching API logic
        total_income = get_income_amount([
            'rental_income', 'water_rates_recovered', 'bank_contribution',
            'rent', 'gross_rent', 'rent_received', 'other_income'
        ])

        # Interest - API uses category_code == 'interest' with deductible_amount
        interest_expense = get_expense_amount([
            'interest', 'interest_expense', 'mortgage_interest', 'loan_interest'
        ])

        # Other deductible expenses (non-interest)
        other_deductible = get_expense_amount([
            'agent_fees', 'property_management_fees', 'property_management',
            'rates', 'council_rates', 'water_rates',
            'insurance', 'landlord_insurance',
            'body_corporate', 'body_corp', 'bc_levies',
            'repairs_maintenance', 'repairs', 'maintenance',
            'consulting_accounting', 'accounting_fees',
            'depreciation', 'bank_fees', 'cleaning', 'advertising',
            'legal_fees', 'letting_fee', 'due_diligence'
        ])

        # Add standard accounting fee if not present
        if not summary_by_cat.get('consulting_accounting') and not summary_by_cat.get('accounting_fees'):
            other_deductible += Decimal('862.50')

        total_deductions = interest_expense + other_deductible
        net_rental_income = total_income - total_deductions

        logger.info("=== WORKBOOK VERIFICATION (matching API logic) ===")
        logger.info(f"Total Income: ${total_income:,.2f}")
        logger.info(f"Interest Expense (deductible): ${interest_expense:,.2f}")
        logger.info(f"Other Deductible: ${other_deductible:,.2f}")
        logger.info(f"Total Deductions: ${total_deductions:,.2f}")
        logger.info(f"Net Rental Income: ${net_rental_income:,.2f}")

        return {
            'total_income': float(total_income),
            'interest_expense': float(interest_expense),
            'other_deductible': float(other_deductible),
            'total_deductions': float(total_deductions),
            'net_rental_income': float(net_rental_income),
        }

    def _build_profit_loss_sheet_from_workings(self, ws: Worksheet, context: Dict[str, Any]):
        """
        Build the Profit and Loss sheet using AI Brain workings data.

        This uses TaxReturnWorkingsData which has accurate calculations from the AI Brain.
        """
        tax_return = context["tax_return"]
        workings: TaxReturnWorkingsData = context["workings"]
        transactions = context["transactions"]
        interest_deductibility = context["interest_deductibility"]

        # Log the workings data for verification
        logger.info("=== AI BRAIN WORKINGS DATA ===")
        logger.info(f"Summary: income={workings.summary.total_income}, deductions={workings.summary.total_deductions}")
        logger.info(f"Interest: gross={workings.summary.interest_gross}, deductible={workings.summary.interest_deductible_amount}")
        logger.info(f"Net Rental Income: {workings.summary.net_rental_income}")

        # Set column widths (exact match to template)
        ws.column_dimensions["A"].width = 29.33
        ws.column_dimensions["B"].width = 4.16
        ws.column_dimensions["C"].width = 12.33
        ws.column_dimensions["D"].width = 13.0
        ws.column_dimensions["E"].width = 7.0
        ws.column_dimensions["F"].width = 12.5
        ws.column_dimensions["G"].width = 12.5
        ws.column_dimensions["H"].width = 2.0  # Spacer
        ws.column_dimensions["I"].width = 19.83
        ws.column_dimensions["J"].width = 4.16
        ws.column_dimensions["K"].width = 13.5
        ws.column_dimensions["L"].width = 13.5
        ws.column_dimensions["M"].width = 13.5
        ws.column_dimensions["N"].width = 13.5
        ws.column_dimensions["O"].width = 11.33
        ws.column_dimensions["P"].width = 11.5
        ws.column_dimensions["Q"].width = 10.0
        ws.column_dimensions["R"].width = 15.0

        # =====================================================================
        # LEFT SIDE - P&L SUMMARY (from AI Brain workings)
        # =====================================================================

        # Row 1: Client name and Business Commencement Date
        ws["A1"] = tax_return.client.name
        ws["A1"].font = Font(bold=True)
        ws["C1"] = "Business Commencement Date - XX/XX/XXXX"

        # Row 2: Resident for Tax Purposes
        ws["A2"] = "Resident for Tax Purposes"
        ws["B2"] = "Y"
        ws["B2"].alignment = Alignment(horizontal="center")

        # Row 3: Property Ownership headers
        ws["C3"] = "Property Ownership"
        ws["F3"] = "Property Ownership"

        # Row 4: Property addresses
        ws["C4"] = tax_return.property_address
        ws["F4"] = "Property Name"

        # Row 5: Ownership percentage
        ws["C5"] = 1
        ws["C5"].number_format = PERCENTAGE_FORMAT
        ws["F5"] = 1
        ws["F5"].number_format = PERCENTAGE_FORMAT

        # === INCOME SECTION (from workings.income) ===
        def write_income_line(row: int, label: str, line_item: Optional[LineItem]):
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            if line_item and line_item.gross_amount:
                ws[f"B{row}"] = line_item.source_code
                ws[f"C{row}"] = float(abs(line_item.gross_amount))
                logger.info(f"Income {label}: ${float(abs(line_item.gross_amount)):,.2f} (source: {line_item.source_code})")
            else:
                ws[f"C{row}"] = 0
            ws[f"C{row}"].number_format = CURRENCY_FORMAT

        write_income_line(6, "Rental Income", workings.income.rental_income)
        write_income_line(7, "Water Rates Recovered", workings.income.water_rates_recovered)
        write_income_line(8, "Bank Contribution", workings.income.bank_contribution)

        # Row 9: Total Income
        ws["A9"] = "Total Income"
        ws["A9"].font = Font(bold=True)
        ws["A9"].alignment = Alignment(horizontal="center")
        ws["C9"] = "=SUM(C6:C8)"
        ws["C9"].number_format = CURRENCY_FORMAT
        ws["F9"] = "=SUM(F6:F8)"
        ws["F9"].number_format = CURRENCY_FORMAT

        # Row 11: Expenses header
        ws["A11"] = "Expenses"
        ws["A11"].font = Font(bold=True)
        ws["A11"].alignment = Alignment(horizontal="center")

        # === EXPENSE SECTION (from workings.expenses) ===
        def write_expense_line(row: int, label: str, line_item: Optional[LineItem], show_percentage: bool = False):
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            if line_item:
                # Use deductible_amount for P&L (what can be claimed)
                amount = line_item.deductible_amount if line_item.deductible_amount else line_item.gross_amount
                if amount:
                    ws[f"B{row}"] = line_item.source_code
                    ws[f"C{row}"] = float(abs(amount))
                    if show_percentage and line_item.deductible_percentage < 100:
                        ws[f"E{row}"] = f"{line_item.deductible_percentage:.0f}%"
                    logger.info(f"Expense {label}: ${float(abs(amount)):,.2f} (source: {line_item.source_code})")
                else:
                    ws[f"C{row}"] = 0
            else:
                ws[f"C{row}"] = 0
            ws[f"C{row}"].number_format = CURRENCY_FORMAT

        # Expense lines from workings
        write_expense_line(12, "Advertising", workings.expenses.advertising)
        write_expense_line(13, "Agent Fees", workings.expenses.agent_fees)
        ws["A14"] = "Assets Under $500"
        ws["A14"].font = Font(bold=True)
        ws["C14"] = 0
        ws["C14"].number_format = CURRENCY_FORMAT
        write_expense_line(15, "Bank Fees", workings.expenses.bank_fees)
        ws["A16"] = "Cleaning"
        ws["A16"].font = Font(bold=True)
        ws["C16"] = 0
        ws["C16"].number_format = CURRENCY_FORMAT

        # Accounting fees - use workings or default
        ws["A17"] = "Consulting & Accounting"
        ws["A17"].font = Font(bold=True)
        if workings.expenses.accounting_fees and workings.expenses.accounting_fees.deductible_amount:
            ws["B17"] = "AF"
            ws["C17"] = float(abs(workings.expenses.accounting_fees.deductible_amount))
        else:
            ws["B17"] = "AF"
            ws["C17"] = 862.50
        ws["C17"].number_format = CURRENCY_FORMAT

        write_expense_line(18, "Depreciation", workings.expenses.depreciation)
        write_expense_line(19, "Due Diligence", workings.expenses.due_diligence)
        ws["A20"] = "Entertainment"
        ws["A20"].font = Font(bold=True)
        ws["C20"] = 0
        ws["C20"].number_format = CURRENCY_FORMAT
        ws["A21"] = "Entertainment - Non deductible"
        ws["A21"].font = Font(bold=True)
        ws["C21"] = 0
        ws["C21"].number_format = CURRENCY_FORMAT
        ws["A22"] = "Freight & Courier"
        ws["A22"].font = Font(bold=True)
        ws["C22"] = 0
        ws["C22"].number_format = CURRENCY_FORMAT
        write_expense_line(23, "General Expenses", workings.expenses.other_expenses)
        ws["A24"] = "Home Office Expense"
        ws["A24"].font = Font(bold=True)
        ws["C24"] = 0
        ws["C24"].number_format = CURRENCY_FORMAT
        write_expense_line(25, "Insurance", workings.expenses.insurance)

        # Interest - show deductible amount with percentage
        write_expense_line(26, "Interest Expense", workings.expenses.interest, show_percentage=True)

        write_expense_line(27, "Legal Expenses", workings.expenses.legal_fees)
        ws["A28"] = "Light, Power, Heating"
        ws["A28"].font = Font(bold=True)
        ws["C28"] = 0
        ws["C28"].number_format = CURRENCY_FORMAT
        ws["A29"] = "Loss on Disposal of Fixed Asset"
        ws["A29"].font = Font(bold=True)
        ws["C29"] = 0
        ws["C29"].number_format = CURRENCY_FORMAT
        ws["A30"] = "Motor Vehicle Expenses"
        ws["A30"].font = Font(bold=True)
        ws["C30"] = 0
        ws["C30"].number_format = CURRENCY_FORMAT
        ws["A31"] = "Office Expenses"
        ws["A31"].font = Font(bold=True)
        ws["C31"] = 0
        ws["C31"].number_format = CURRENCY_FORMAT
        ws["A32"] = "Overdraft Interest"
        ws["A32"].font = Font(bold=True)
        ws["C32"] = 0
        ws["C32"].number_format = CURRENCY_FORMAT
        ws["A33"] = "Printing & Stationery"
        ws["A33"].font = Font(bold=True)
        ws["C33"] = 0
        ws["C33"].number_format = CURRENCY_FORMAT
        write_expense_line(34, "Rates", workings.expenses.rates)
        write_expense_line(35, "Repairs & Maintenance", workings.expenses.repairs_maintenance)
        ws["A36"] = "Shareholder Salary"
        ws["A36"].font = Font(bold=True)
        ws["C36"] = 0
        ws["C36"].number_format = CURRENCY_FORMAT
        ws["A37"] = "Subscriptions"
        ws["A37"].font = Font(bold=True)
        ws["C37"] = 0
        ws["C37"].number_format = CURRENCY_FORMAT
        ws["A38"] = "Telephone & Internet"
        ws["A38"].font = Font(bold=True)
        ws["C38"] = 0
        ws["C38"].number_format = CURRENCY_FORMAT
        ws["A39"] = "Travel - National"
        ws["A39"].font = Font(bold=True)
        ws["C39"] = 0
        ws["C39"].number_format = CURRENCY_FORMAT
        ws["A40"] = "Travel - International"
        ws["A40"].font = Font(bold=True)
        ws["C40"] = 0
        ws["C40"].number_format = CURRENCY_FORMAT
        write_expense_line(41, "Water Rates", workings.expenses.water_rates)
        write_expense_line(42, "Body Corporate", workings.expenses.body_corporate)

        # Row 43: Total Expenses
        ws["A43"] = "Total Expenses"
        ws["A43"].font = Font(bold=True)
        ws["A43"].alignment = Alignment(horizontal="center")
        ws["C43"] = "=SUM(C12:C42)"
        ws["C43"].number_format = CURRENCY_FORMAT
        ws["F43"] = "=SUM(F12:F42)"
        ws["F43"].number_format = CURRENCY_FORMAT

        # Row 44: Net Income
        ws["A44"] = "Net Income"
        ws["A44"].font = Font(bold=True)
        ws["A44"].alignment = Alignment(horizontal="center")
        ws["C44"] = "=C9-C43"
        ws["C44"].number_format = CURRENCY_FORMAT
        ws["F44"] = "=F9-F43"
        ws["F44"].number_format = CURRENCY_FORMAT

        # Row 46: Add back rental profit/loss
        ws["A46"] = "Add back rental profit/loss (EL 4 ITA 2007)"

        # =====================================================================
        # RIGHT SIDE - WORKINGS (same as before, from transactions)
        # =====================================================================
        self._build_workings_section(ws, context)

    def _build_workings_section(self, ws: Worksheet, context: Dict[str, Any]):
        """Build the right side workings section of the P&L sheet."""
        tax_return = context["tax_return"]
        transactions = context["transactions"]
        interest_deductibility = context["interest_deductibility"]
        deductibility_rate = context["deductibility_rate"]

        # Row 1: Additional Information headers
        ws["I1"] = "Additional Information"
        ws["O1"] = "IRD Look Up"

        # Row 2: IRD Look Up headers
        ws["P2"] = "Unfiled"
        ws["Q2"] = "Amount Due"
        ws["R2"] = "Notes"

        # Row 3: Other Income section
        ws["I3"] = "Other Income"
        ws["K3"] = "Gross"
        ws["L3"] = "WT"
        ws["O3"] = "GST"

        # Row 11: Interest Earnings section
        ws["I11"] = "Interest Earnings"
        ws["K11"] = "Gross"
        ws["L11"] = "RWT"

        # Row 12: Bank Name
        ws["I12"] = "Bank Name"
        ws["J12"] = "B/S"

        # Row 16: Dividends section
        ws["I16"] = "Dividends"
        ws["K16"] = "Gross"
        ws["L16"] = "Imputation"
        ws["M16"] = "RWT"

        # Row 21: Donations
        ws["I21"] = "Donations"

        # Row 25: Excess Residential Deductions
        ws["I25"] = "Excess Residential Deductions Carried Forward"

        # Row 27: Client name reference
        ws["I27"] = "=A1"

        # Row 32: Interest Deductibility and Bank Statement Workings
        ws["I32"] = "Interest Deductibility and Bank Statement Workings"
        ws["I32"].font = Font(bold=True)

        # Extract loan accounts and get monthly data
        loan_accounts = self._extract_loan_accounts(transactions)
        monthly_interest = self._group_interest_by_month(transactions, tax_return.tax_year)
        monthly_other = self._group_other_by_month(transactions, tax_return.tax_year)

        # Row 33: Column headers
        for i, loan_name in enumerate(loan_accounts[:3]):
            ws.cell(row=33, column=11+i, value=loan_name)
        ws["N33"] = "Rates"
        ws["O33"] = "Insurance"
        ws["P33"] = "Bank Fees"

        # Rows 34-45: Monthly data
        fy_months = self._get_fy_months(tax_return.tax_year)
        for idx, (month_dt, month_key) in enumerate(fy_months):
            row_num = 34 + idx
            ws.cell(row=row_num, column=9, value=month_dt.strftime("%b-%y"))
            ws.cell(row=row_num, column=10, value="BS")

            if month_key in monthly_interest:
                for i, loan_name in enumerate(loan_accounts[:3]):
                    if loan_name in monthly_interest[month_key]:
                        cell = ws.cell(row=row_num, column=11+i, value=float(monthly_interest[month_key][loan_name]))
                        cell.number_format = CURRENCY_FORMAT

            if month_key in monthly_other and "rates" in monthly_other[month_key]:
                ws.cell(row=row_num, column=14, value=float(monthly_other[month_key]["rates"]))
                ws.cell(row=row_num, column=14).number_format = CURRENCY_FORMAT

            if month_key in monthly_other and "insurance" in monthly_other[month_key]:
                ws.cell(row=row_num, column=15, value=float(monthly_other[month_key]["insurance"]))
                ws.cell(row=row_num, column=15).number_format = CURRENCY_FORMAT

            if month_key in monthly_other and "bank_fees" in monthly_other[month_key]:
                ws.cell(row=row_num, column=16, value=float(monthly_other[month_key]["bank_fees"]))
                ws.cell(row=row_num, column=16).number_format = CURRENCY_FORMAT

        # Row 46: Totals
        ws["I46"] = "Total"
        ws["K46"] = "=SUM(K34:K45)"
        ws["K46"].number_format = CURRENCY_FORMAT
        ws["L46"] = "=SUM(L34:L45)"
        ws["L46"].number_format = CURRENCY_FORMAT
        ws["M46"] = "=SUM(M34:M45)"
        ws["M46"].number_format = CURRENCY_FORMAT
        ws["N46"] = "=SUM(N34:N45)"
        ws["N46"].number_format = CURRENCY_FORMAT
        ws["O46"] = "=SUM(O34:O45)"
        ws["O46"].number_format = CURRENCY_FORMAT
        ws["P46"] = "=SUM(P34:P45)"
        ws["P46"].number_format = CURRENCY_FORMAT

        # Row 47: Deductible amounts
        ws["I47"] = f"Deductible ({interest_deductibility}%)"
        ws["K47"] = f"=K46*{deductibility_rate}"
        ws["K47"].number_format = CURRENCY_FORMAT
        ws["L47"] = f"=L46*{deductibility_rate}"
        ws["L47"].number_format = CURRENCY_FORMAT
        ws["M47"] = f"=M46*{deductibility_rate}"
        ws["M47"].number_format = CURRENCY_FORMAT

        # Row 48: Capitalised Interest
        ws["I48"] = "Capitalised Interest"
        ws["K48"] = "=K46-K47"
        ws["K48"].number_format = CURRENCY_FORMAT
        ws["L48"] = "=L46-L47"
        ws["L48"].number_format = CURRENCY_FORMAT
        ws["M48"] = "=M46-M47"
        ws["M48"].number_format = CURRENCY_FORMAT

        # Bottom sections (repairs, capital, PM statements)
        self._build_bottom_sections(ws, context)

    def _build_bottom_sections(self, ws: Worksheet, context: Dict[str, Any]):
        """Build the bottom sections of the P&L sheet."""
        tax_return = context["tax_return"]
        transactions = context["transactions"]

        # Row 49: Additional Information
        ws["A49"] = "Additional Information"

        # Row 50: Repairs and Maintenance header
        ws["A50"] = "Repairs and Maintenance"
        ws["A50"].font = Font(bold=True)
        ws["C50"] = "Amount"
        ws["D50"] = "Date"
        ws["E50"] = "Invoice"

        # Rows 51-55: Repairs items
        repairs_items = self._get_repairs_items(transactions)
        for i, item in enumerate(repairs_items[:5]):
            row_num = 51 + i
            ws[f"A{row_num}"] = item["description"][:40] if item["description"] else ""
            ws[f"C{row_num}"] = float(abs(item["amount"]))
            ws[f"C{row_num}"].number_format = CURRENCY_FORMAT
            if item["date"]:
                ws[f"D{row_num}"] = item["date"].strftime("%d/%m/%Y") if hasattr(item["date"], 'strftime') else str(item["date"])
            ws[f"E{row_num}"] = "Y/N"

        # Row 56: Capital header
        ws["A56"] = "Capital"
        ws["A56"].font = Font(bold=True)
        ws["C56"] = "Amount"
        ws["D56"] = "Date"
        ws["E56"] = "Invoice"

        # Rows 57-60: Capital items
        capital_items = self._get_capital_items(transactions)
        for i, item in enumerate(capital_items[:4]):
            row_num = 57 + i
            ws[f"A{row_num}"] = item["description"][:40] if item["description"] else ""
            ws[f"C{row_num}"] = float(abs(item["amount"]))
            ws[f"C{row_num}"].number_format = CURRENCY_FORMAT
            if item["date"]:
                ws[f"D{row_num}"] = item["date"].strftime("%d/%m/%Y") if hasattr(item["date"], 'strftime') else str(item["date"])
            ws[f"E{row_num}"] = "Y/N"

        # Row 64: Notes
        ws["A64"] = "Notes:"
        ws["A64"].font = Font(bold=True)

        # Rows 70-78: Information Source Key
        ws["A70"] = "Information Source Key"
        ws["A70"].font = Font(bold=True)

        source_codes = [
            (71, "Additional Information", "AI"),
            (72, "Accounting Fee", "AF"),
            (73, "Bank Statement", "BS"),
            (74, "Client Provided", "CP"),
            (75, "Invoice/Receipt", "INV"),
            (76, "Inland Revenue", "IR"),
            (77, "Property Manager", "PM"),
            (78, "Prior Accountant", "PA"),
        ]
        for row_num, description, code in source_codes:
            ws[f"A{row_num}"] = description
            ws[f"B{row_num}"] = code

        # PM Statements section
        ws["I50"] = "Property Manager Statements (use when no Year End Statement)"
        ws["I50"].font = Font(bold=True)
        ws["K51"] = "Rental Income"
        ws["L51"] = "Agent Fees"
        ws["M51"] = "Repairs and Maintenance"

        # Get monthly PM data
        monthly_pm = self._group_pm_by_month(transactions, tax_return.tax_year)
        fy_months = self._get_fy_months(tax_return.tax_year)

        for idx, (month_dt, month_key) in enumerate(fy_months):
            row_num = 52 + idx
            ws.cell(row=row_num, column=9, value=month_dt.strftime("%b-%y"))
            ws.cell(row=row_num, column=10, value="PM")

            if month_key in monthly_pm:
                if "rental_income" in monthly_pm[month_key]:
                    ws.cell(row=row_num, column=11, value=float(monthly_pm[month_key]["rental_income"]))
                    ws.cell(row=row_num, column=11).number_format = CURRENCY_FORMAT
                if "agent_fees" in monthly_pm[month_key]:
                    ws.cell(row=row_num, column=12, value=float(monthly_pm[month_key]["agent_fees"]))
                    ws.cell(row=row_num, column=12).number_format = CURRENCY_FORMAT
                if "repairs_maintenance" in monthly_pm[month_key]:
                    ws.cell(row=row_num, column=13, value=float(monthly_pm[month_key]["repairs_maintenance"]))
                    ws.cell(row=row_num, column=13).number_format = CURRENCY_FORMAT

        # Row 64: PM Totals
        ws["K64"] = "=SUM(K52:K63)"
        ws["K64"].number_format = CURRENCY_FORMAT
        ws["L64"] = "=SUM(L52:L63)"
        ws["L64"].number_format = CURRENCY_FORMAT
        ws["M64"] = "=SUM(M52:M63)"
        ws["M64"].number_format = CURRENCY_FORMAT

    def _build_profit_loss_sheet(self, ws: Worksheet, context: Dict[str, Any]):
        """Build the Profit and Loss sheet matching Lighthouse template exactly."""
        tax_return = context["tax_return"]
        transactions = context["transactions"]
        summaries = context["summaries"]
        deductibility_rate = context["deductibility_rate"]
        interest_deductibility = context["interest_deductibility"]

        # DEBUG: Log all available category codes and amounts
        logger.info("=== AVAILABLE SUMMARIES ===")
        for s in summaries:
            logger.info(f"  {s.category_code}: gross={s.gross_amount}, deductible={s.deductible_amount}, count={s.transaction_count}")

        # Create lookups
        summary_by_category = {s.category_code: s for s in summaries}
        logger.info(f"Category codes available: {list(summary_by_category.keys())}")

        # Check for duplicates
        self._check_for_duplicates(summaries)

        # Verify totals
        self._verify_totals(context)

        # Set column widths (exact match to template)
        ws.column_dimensions["A"].width = 29.33
        ws.column_dimensions["B"].width = 4.16
        ws.column_dimensions["C"].width = 12.33
        ws.column_dimensions["D"].width = 13.0
        ws.column_dimensions["E"].width = 7.0
        ws.column_dimensions["F"].width = 12.5
        ws.column_dimensions["G"].width = 12.5
        ws.column_dimensions["H"].width = 2.0  # Spacer
        ws.column_dimensions["I"].width = 19.83
        ws.column_dimensions["J"].width = 4.16
        ws.column_dimensions["K"].width = 13.5
        ws.column_dimensions["L"].width = 13.5
        ws.column_dimensions["M"].width = 13.5
        ws.column_dimensions["N"].width = 13.5
        ws.column_dimensions["O"].width = 11.33
        ws.column_dimensions["P"].width = 11.5
        ws.column_dimensions["Q"].width = 10.0
        ws.column_dimensions["R"].width = 15.0

        # =====================================================================
        # LEFT SIDE - P&L SUMMARY
        # =====================================================================

        # Row 1: Client name and Business Commencement Date
        ws["A1"] = tax_return.client.name
        ws["A1"].font = Font(bold=True)
        ws["C1"] = "Business Commencement Date - XX/XX/XXXX"

        # Row 2: Resident for Tax Purposes
        ws["A2"] = "Resident for Tax Purposes"
        ws["B2"] = "Y"
        ws["B2"].alignment = Alignment(horizontal="center")

        # Row 3: Property Ownership headers
        ws["C3"] = "Property Ownership"
        ws["F3"] = "Property Ownership"

        # Row 4: Property addresses
        ws["C4"] = tax_return.property_address
        ws["F4"] = "Property Name"

        # Row 5: Ownership percentage (as decimal with 0% format)
        ws["C5"] = 1
        ws["C5"].number_format = PERCENTAGE_FORMAT
        ws["F5"] = 1
        ws["F5"].number_format = PERCENTAGE_FORMAT

        # === INCOME SECTION ===
        # Income lines: (row, label, primary_code, [alt_codes])
        # NOTE: For income, we use gross_amount (total rent received)
        # The API uses: sum(t.amount for t in transactions if t.amount > 0 and t.category_code in income_categories)
        income_lines = [
            (6, "Rental Income", "rental_income", ["rent", "gross_rent", "rent_received"]),
            (7, "Water Rates Recovered", "water_rates_recovered", ["water_recovered"]),
            (8, "Bank Contribution", "bank_contribution", ["bank_contrib", "interest_income"]),
        ]

        for row_num, label, primary_code, alt_codes in income_lines:
            ws[f"A{row_num}"] = label
            ws[f"A{row_num}"].font = Font(bold=True)

            summary = self._get_summary(summary_by_category, primary_code, alt_codes)
            if summary:
                # For income, use gross_amount (positive amounts)
                amount = summary.gross_amount
                if amount:
                    ws[f"B{row_num}"] = self._get_source_code(summary, primary_code)
                    ws[f"C{row_num}"] = float(abs(amount))
                    logger.info(f"Income {primary_code}: ${float(abs(amount)):,.2f}")
                else:
                    ws[f"C{row_num}"] = 0
            else:
                ws[f"C{row_num}"] = 0
            ws[f"C{row_num}"].number_format = CURRENCY_FORMAT

        # Row 9: Total Income
        ws["A9"] = "Total Income"
        ws["A9"].font = Font(bold=True)
        ws["A9"].alignment = Alignment(horizontal="center")
        ws["C9"] = "=SUM(C6:C8)"
        ws["C9"].number_format = CURRENCY_FORMAT
        ws["F9"] = "=SUM(F6:F8)"
        ws["F9"].number_format = CURRENCY_FORMAT

        # Row 11: Expenses header
        ws["A11"] = "Expenses"
        ws["A11"].font = Font(bold=True)
        ws["A11"].alignment = Alignment(horizontal="center")

        # === EXPENSE SECTION ===
        # Expense lines: (row, label, primary_code, [alt_codes])
        # CRITICAL: For expenses, we must use deductible_amount, not gross_amount
        # The API uses: sum(abs(t.deductible_amount if t.deductible_amount is not None else t.amount) ...)
        expense_lines = [
            (12, "Advertising", "advertising", ["ad_costs", "marketing"]),
            (13, "Agent Fees", "agent_fees", ["property_management_fees", "property_management", "pm_fees", "letting_fee"]),
            (14, "Assets Under $500", "assets_under_500", ["minor_assets", "small_assets"]),
            (15, "Bank Fees", "bank_fees", ["bank_charges"]),
            (16, "Cleaning", "cleaning", ["cleaning_expenses"]),
            (17, "Consulting & Accounting", "consulting_accounting", ["accounting_fees", "accountant"]),
            (18, "Depreciation", "depreciation", ["depreciation_expense"]),
            (19, "Due Diligence", "due_diligence", ["lim_report", "meth_test", "valuations"]),
            (20, "Entertainment", "entertainment", []),
            (21, "Entertainment - Non deductible", "entertainment_non_deductible", []),
            (22, "Freight & Courier", "freight_courier", ["courier", "freight"]),
            (23, "General Expenses", "general_expenses", ["other_expenses", "miscellaneous"]),
            (24, "Home Office Expense", "home_office", ["home_office_expense"]),
            (25, "Insurance", "insurance", ["landlord_insurance", "property_insurance"]),
            # Row 26: Interest - NOTE: API uses category_code == 'interest' (not 'interest_expense')
            (26, "Interest Expense", "interest", ["interest_expense", "mortgage_interest", "loan_interest"]),
            (27, "Legal Expenses", "legal_fees", ["legal_expenses", "solicitor"]),
            (28, "Light, Power, Heating", "electricity", ["power", "utilities"]),
            (29, "Loss on Disposal of Fixed Asset", "loss_on_disposal", []),
            (30, "Motor Vehicle Expenses", "vehicle_expenses", ["motor_vehicle", "mileage"]),
            (31, "Office Expenses", "office_expenses", ["office_supplies"]),
            (32, "Overdraft Interest", "overdraft_interest", []),
            (33, "Printing & Stationery", "printing_stationery", ["stationery"]),
            (34, "Rates", "rates", ["council_rates", "local_rates"]),
            (35, "Repairs & Maintenance", "repairs_maintenance", ["repairs", "maintenance"]),
            (36, "Shareholder Salary", "shareholder_salary", []),
            (37, "Subscriptions", "subscriptions", []),
            (38, "Telephone & Internet", "telephone_internet", ["phone", "internet"]),
            (39, "Travel - National", "travel_national", ["domestic_travel"]),
            (40, "Travel - International", "travel_international", ["overseas_travel"]),
            (41, "Water Rates", "water_rates", ["water_charges"]),
            (42, "Body Corporate", "body_corporate", ["body_corp", "bc_levies", "strata_fees"]),
        ]

        for row_num, label, primary_code, alt_codes in expense_lines:
            ws[f"A{row_num}"] = label
            ws[f"A{row_num}"].font = Font(bold=True)

            # Special handling for certain categories
            if primary_code == "consulting_accounting":
                # Fixed accounting fee - check if already in summaries first
                summary = self._get_summary(summary_by_category, primary_code, alt_codes)
                if summary:
                    amount = summary.deductible_amount if summary.deductible_amount else summary.gross_amount
                    if amount:
                        ws[f"B{row_num}"] = "AF"
                        ws[f"C{row_num}"] = float(abs(amount))
                    else:
                        ws[f"B{row_num}"] = "AF"
                        ws[f"C{row_num}"] = 862.50
                else:
                    # Default accounting fee
                    ws[f"B{row_num}"] = "AF"
                    ws[f"C{row_num}"] = 862.50

            elif primary_code == "interest":
                # Interest Expense - MUST use deductible_amount (already has 80% applied)
                # The API uses: sum(abs(t.deductible_amount if t.deductible_amount is not None else t.amount)
                #               for t in transactions if t.category_code == 'interest' and t.amount < 0)
                summary = self._get_summary(summary_by_category, primary_code, alt_codes)
                if summary:
                    # Use deductible_amount if available (should already have deductibility applied)
                    if summary.deductible_amount:
                        deductible_interest = float(abs(summary.deductible_amount))
                        ws[f"B{row_num}"] = "BS"
                        ws[f"C{row_num}"] = deductible_interest
                        ws[f"E{row_num}"] = f"{interest_deductibility}%"
                        logger.info(f"Interest from deductible_amount: ${deductible_interest:,.2f}")
                    elif summary.gross_amount:
                        # Fallback: apply deductibility rate to gross
                        total_interest = float(abs(summary.gross_amount))
                        deductible_interest = total_interest * deductibility_rate
                        ws[f"B{row_num}"] = "BS"
                        ws[f"C{row_num}"] = deductible_interest
                        ws[f"E{row_num}"] = f"{interest_deductibility}%"
                        logger.info(f"Interest calculated: ${total_interest:,.2f} x {deductibility_rate} = ${deductible_interest:,.2f}")
                    else:
                        ws[f"C{row_num}"] = 0
                else:
                    ws[f"C{row_num}"] = 0
                    logger.warning("No interest summary found!")
            else:
                # Standard expense lookup - USE DEDUCTIBLE AMOUNT
                summary = self._get_summary(summary_by_category, primary_code, alt_codes)
                if summary:
                    # CRITICAL: Use deductible_amount for P&L (what can be claimed as deduction)
                    # Fall back to gross_amount only if deductible_amount is not set
                    amount = summary.deductible_amount if summary.deductible_amount else summary.gross_amount
                    if amount:
                        ws[f"B{row_num}"] = self._get_source_code(summary, primary_code)
                        ws[f"C{row_num}"] = float(abs(amount))
                        logger.info(f"Expense {primary_code}: deductible=${float(abs(amount)):,.2f} (gross=${float(abs(summary.gross_amount or 0)):,.2f})")
                    else:
                        ws[f"C{row_num}"] = 0
                else:
                    ws[f"C{row_num}"] = 0

            ws[f"C{row_num}"].number_format = CURRENCY_FORMAT

        # Row 43: Total Expenses
        ws["A43"] = "Total Expenses"
        ws["A43"].font = Font(bold=True)
        ws["A43"].alignment = Alignment(horizontal="center")
        ws["C43"] = "=SUM(C12:C42)"
        ws["C43"].number_format = CURRENCY_FORMAT
        ws["F43"] = "=SUM(F12:F42)"
        ws["F43"].number_format = CURRENCY_FORMAT

        # Row 44: Net Income
        ws["A44"] = "Net Income"
        ws["A44"].font = Font(bold=True)
        ws["A44"].alignment = Alignment(horizontal="center")
        ws["C44"] = "=C9-C43"
        ws["C44"].number_format = CURRENCY_FORMAT
        ws["F44"] = "=F9-F43"
        ws["F44"].number_format = CURRENCY_FORMAT

        # Row 46: Add back rental profit/loss
        ws["A46"] = "Add back rental profit/loss (EL 4 ITA 2007)"

        # =====================================================================
        # RIGHT SIDE - WORKINGS
        # =====================================================================

        # Row 1: Additional Information headers
        ws["I1"] = "Additional Information"
        ws["O1"] = "IRD Look Up"

        # Row 2: IRD Look Up headers
        ws["P2"] = "Unfiled"
        ws["Q2"] = "Amount Due"
        ws["R2"] = "Notes"

        # Row 3: Other Income section
        ws["I3"] = "Other Income"
        ws["K3"] = "Gross"
        ws["L3"] = "WT"
        ws["O3"] = "GST"

        # Row 11: Interest Earnings section
        ws["I11"] = "Interest Earnings"
        ws["K11"] = "Gross"
        ws["L11"] = "RWT"

        # Row 12: Bank Name
        ws["I12"] = "Bank Name"
        ws["J12"] = "B/S"

        # Row 16: Dividends section
        ws["I16"] = "Dividends"
        ws["K16"] = "Gross"
        ws["L16"] = "Imputation"
        ws["M16"] = "RWT"

        # Row 21: Donations
        ws["I21"] = "Donations"

        # Row 25: Excess Residential Deductions
        ws["I25"] = "Excess Residential Deductions Carried Forward"

        # Row 27: Client name reference
        ws["I27"] = "=A1"

        # Row 32: Interest Deductibility and Bank Statement Workings
        ws["I32"] = "Interest Deductibility and Bank Statement Workings"
        ws["I32"].font = Font(bold=True)

        # Extract loan accounts and get monthly data
        loan_accounts = self._extract_loan_accounts(transactions)
        monthly_interest = self._group_interest_by_month(transactions, tax_return.tax_year)
        monthly_other = self._group_other_by_month(transactions, tax_return.tax_year)

        # Row 33: Column headers
        for i, loan_name in enumerate(loan_accounts[:3]):
            ws.cell(row=33, column=11+i, value=loan_name)
        ws["N33"] = "Rates"
        ws["O33"] = "Insurance"
        ws["P33"] = "Bank Fees"

        # Rows 34-45: Monthly data
        fy_months = self._get_fy_months(tax_return.tax_year)
        for idx, (month_dt, month_key) in enumerate(fy_months):
            row_num = 34 + idx

            # Month label - format as mmm-yy
            ws.cell(row=row_num, column=9, value=month_dt.strftime("%b-%y"))
            ws.cell(row=row_num, column=10, value="BS")

            # Interest by loan account
            if month_key in monthly_interest:
                for i, loan_name in enumerate(loan_accounts[:3]):
                    if loan_name in monthly_interest[month_key]:
                        cell = ws.cell(row=row_num, column=11+i, value=float(monthly_interest[month_key][loan_name]))
                        cell.number_format = CURRENCY_FORMAT

            # Rates (column N)
            if month_key in monthly_other and "rates" in monthly_other[month_key]:
                ws.cell(row=row_num, column=14, value=float(monthly_other[month_key]["rates"]))
                ws.cell(row=row_num, column=14).number_format = CURRENCY_FORMAT

            # Insurance (column O)
            if month_key in monthly_other and "insurance" in monthly_other[month_key]:
                ws.cell(row=row_num, column=15, value=float(monthly_other[month_key]["insurance"]))
                ws.cell(row=row_num, column=15).number_format = CURRENCY_FORMAT

            # Bank Fees (column P)
            if month_key in monthly_other and "bank_fees" in monthly_other[month_key]:
                ws.cell(row=row_num, column=16, value=float(monthly_other[month_key]["bank_fees"]))
                ws.cell(row=row_num, column=16).number_format = CURRENCY_FORMAT

        # Row 46: Totals (Gross interest)
        ws["I46"] = "Total"
        ws["K46"] = "=SUM(K34:K45)"
        ws["K46"].number_format = CURRENCY_FORMAT
        ws["L46"] = "=SUM(L34:L45)"
        ws["L46"].number_format = CURRENCY_FORMAT
        ws["M46"] = "=SUM(M34:M45)"
        ws["M46"].number_format = CURRENCY_FORMAT
        ws["N46"] = "=SUM(N34:N45)"
        ws["N46"].number_format = CURRENCY_FORMAT
        ws["O46"] = "=SUM(O34:O45)"
        ws["O46"].number_format = CURRENCY_FORMAT
        ws["P46"] = "=SUM(P34:P45)"
        ws["P46"].number_format = CURRENCY_FORMAT

        # Row 47: Deductible amounts
        ws["I47"] = f"Deductible ({interest_deductibility}%)"
        ws["K47"] = f"=K46*{deductibility_rate}"
        ws["K47"].number_format = CURRENCY_FORMAT
        ws["L47"] = f"=L46*{deductibility_rate}"
        ws["L47"].number_format = CURRENCY_FORMAT
        ws["M47"] = f"=M46*{deductibility_rate}"
        ws["M47"].number_format = CURRENCY_FORMAT

        # Row 48: Capitalised Interest
        ws["I48"] = "Capitalised Interest"
        ws["K48"] = "=K46-K47"
        ws["K48"].number_format = CURRENCY_FORMAT
        ws["L48"] = "=L46-L47"
        ws["L48"].number_format = CURRENCY_FORMAT
        ws["M48"] = "=M46-M47"
        ws["M48"].number_format = CURRENCY_FORMAT

        # =====================================================================
        # BOTTOM LEFT - DETAIL SECTIONS
        # =====================================================================

        # Row 49: Additional Information
        ws["A49"] = "Additional Information"

        # Row 50: Repairs and Maintenance header
        ws["A50"] = "Repairs and Maintenance"
        ws["A50"].font = Font(bold=True)
        ws["C50"] = "Amount"
        ws["D50"] = "Date"
        ws["E50"] = "Invoice"

        # Rows 51-55: Repairs items
        repairs_items = self._get_repairs_items(transactions)
        for i, item in enumerate(repairs_items[:5]):
            row_num = 51 + i
            ws[f"A{row_num}"] = item["description"][:40] if item["description"] else ""
            ws[f"C{row_num}"] = float(abs(item["amount"]))
            ws[f"C{row_num}"].number_format = CURRENCY_FORMAT
            if item["date"]:
                ws[f"D{row_num}"] = item["date"].strftime("%d/%m/%Y") if hasattr(item["date"], 'strftime') else str(item["date"])
            ws[f"E{row_num}"] = "Y/N"

        # Row 56: Capital header
        ws["A56"] = "Capital"
        ws["A56"].font = Font(bold=True)
        ws["C56"] = "Amount"
        ws["D56"] = "Date"
        ws["E56"] = "Invoice"

        # Rows 57-60: Capital items
        capital_items = self._get_capital_items(transactions)
        for i, item in enumerate(capital_items[:4]):
            row_num = 57 + i
            ws[f"A{row_num}"] = item["description"][:40] if item["description"] else ""
            ws[f"C{row_num}"] = float(abs(item["amount"]))
            ws[f"C{row_num}"].number_format = CURRENCY_FORMAT
            if item["date"]:
                ws[f"D{row_num}"] = item["date"].strftime("%d/%m/%Y") if hasattr(item["date"], 'strftime') else str(item["date"])
            ws[f"E{row_num}"] = "Y/N"

        # Row 64: Notes
        ws["A64"] = "Notes:"
        ws["A64"].font = Font(bold=True)

        # Rows 70-78: Information Source Key
        ws["A70"] = "Information Source Key"
        ws["A70"].font = Font(bold=True)

        source_codes = [
            (71, "Additional Information", "AI"),
            (72, "Accounting Fee", "AF"),
            (73, "Bank Statement", "BS"),
            (74, "Client Provided", "CP"),
            (75, "Invoice/Receipt", "INV"),
            (76, "Inland Revenue", "IR"),
            (77, "Property Manager", "PM"),
            (78, "Prior Accountant", "PA"),
        ]
        for row_num, description, code in source_codes:
            ws[f"A{row_num}"] = description
            ws[f"B{row_num}"] = code

        # =====================================================================
        # BOTTOM RIGHT - PROPERTY MANAGER STATEMENTS
        # =====================================================================

        # Row 50: PM Statements header
        ws["I50"] = "Property Manager Statements (use when no Year End Statement)"
        ws["I50"].font = Font(bold=True)

        # Row 51: PM column headers
        ws["K51"] = "Rental Income"
        ws["L51"] = "Agent Fees"
        ws["M51"] = "Repairs and Maintenance"

        # Get monthly PM data
        monthly_pm = self._group_pm_by_month(transactions, tax_return.tax_year)

        # Rows 52-63: Monthly PM data
        for idx, (month_dt, month_key) in enumerate(fy_months):
            row_num = 52 + idx

            ws.cell(row=row_num, column=9, value=month_dt.strftime("%b-%y"))
            ws.cell(row=row_num, column=10, value="PM")

            if month_key in monthly_pm:
                if "rental_income" in monthly_pm[month_key]:
                    ws.cell(row=row_num, column=11, value=float(monthly_pm[month_key]["rental_income"]))
                    ws.cell(row=row_num, column=11).number_format = CURRENCY_FORMAT
                if "agent_fees" in monthly_pm[month_key]:
                    ws.cell(row=row_num, column=12, value=float(monthly_pm[month_key]["agent_fees"]))
                    ws.cell(row=row_num, column=12).number_format = CURRENCY_FORMAT
                if "repairs_maintenance" in monthly_pm[month_key]:
                    ws.cell(row=row_num, column=13, value=float(monthly_pm[month_key]["repairs_maintenance"]))
                    ws.cell(row=row_num, column=13).number_format = CURRENCY_FORMAT

        # Row 64: PM Totals
        ws["K64"] = "=SUM(K52:K63)"
        ws["K64"].number_format = CURRENCY_FORMAT
        ws["L64"] = "=SUM(L52:L63)"
        ws["L64"].number_format = CURRENCY_FORMAT
        ws["M64"] = "=SUM(M52:M63)"
        ws["M64"].number_format = CURRENCY_FORMAT

    def _build_ird_sheet(self, ws: Worksheet, context: Dict[str, Any]):
        """Build the IRD checklist sheet."""

        # Set column widths
        ws.column_dimensions["A"].width = 9.5
        ws.column_dimensions["B"].width = 11.83
        ws.column_dimensions["C"].width = 8.0
        ws.column_dimensions["D"].width = 10.83
        ws.column_dimensions["E"].width = 8.0
        ws.column_dimensions["F"].width = 19.16
        ws.column_dimensions["G"].width = 20.0

        # Row 1: Header
        ws["A1"] = "Have you:"
        ws["A1"].font = Font(bold=True)

        # Row 2: Column headers
        ws["F2"] = "Answer"
        ws["G2"] = "Notes"

        # Questions with merged cells
        questions = [
            (3, "1. Checked WFM/Client File for Notes/Email Correspondence?", "A3:E3"),
            (4, "2. Checked Account Look Up?", "A4:C4"),
            (5, "3. Are there any outstanding Income Tax returns?", "A5:E5"),
            (6, "4. Is there any outstanding Income Tax?", "A6:D6"),
            (7, "5. Are there any outstanding GST returns?", "A7:D7"),
            (8, "6. Is there any outstanding GST?", "A8:C8"),
            (9, "7. Were Provisional Tax Payments made on/before:", "A9:D9"),
        ]

        for row_num, question, merge_range in questions:
            ws[f"A{row_num}"] = question
            ws.merge_cells(merge_range)
            if row_num == 9:
                ws[f"A{row_num}"].alignment = Alignment(horizontal="center")

        # Sub-questions for provisional tax
        sub_questions = [
            (10, "a. 28 August", "A10:B10"),
            (11, "b. 15 January", "A11:B11"),
            (12, "c. 7 May", "A12:B12"),
        ]

        for row_num, question, merge_range in sub_questions:
            ws[f"A{row_num}"] = question
            ws.merge_cells(merge_range)
            ws[f"A{row_num}"].alignment = Alignment(horizontal="center")

    # =========================================================================
    # HELPER METHODS
    # =========================================================================

    def _get_fy_months(self, tax_year: str) -> List[Tuple[datetime, str]]:
        """
        Get list of months for the financial year.

        Returns list of (datetime, "YYYY-MM") tuples for Apr-Mar.
        """
        # Extract year from tax_year (e.g., "FY25" -> 2025)
        year = int("20" + tax_year[-2:])

        months = []
        # Apr to Dec of previous year
        for month in range(4, 13):
            dt = datetime(year - 1, month, 1)
            months.append((dt, dt.strftime("%Y-%m")))
        # Jan to Mar of current year
        for month in range(1, 4):
            dt = datetime(year, month, 1)
            months.append((dt, dt.strftime("%Y-%m")))

        return months

    def _extract_loan_accounts(self, transactions: List[Transaction]) -> List[str]:
        """
        Extract unique loan account names from transaction descriptions.

        Looks for patterns like "91-01", "91-03", "Loan Account 1", etc.
        """
        loan_accounts = set()

        # Patterns for loan account numbers
        patterns = [
            re.compile(r'\b(\d{2}-\d{2})\b'),  # e.g., "91-01", "91-03"
            re.compile(r'Loan\s*(\d+)', re.IGNORECASE),  # e.g., "Loan 1", "Loan2"
            re.compile(r'Account\s*#?\s*(\d+)', re.IGNORECASE),  # e.g., "Account 1"
        ]

        # Categories that indicate interest - 'interest' is the primary code used by API
        interest_categories = {'interest', 'interest_expense', 'mortgage_interest', 'loan_interest'}

        for txn in transactions:
            if txn.category_code in interest_categories and txn.description:
                for pattern in patterns:
                    matches = pattern.findall(txn.description)
                    for match in matches:
                        loan_accounts.add(f"Loan {match}")

        # If no specific accounts found, use generic name
        if not loan_accounts:
            loan_accounts = {"Loan Account 1"}

        return sorted(list(loan_accounts))[:3]  # Max 3 loan accounts

    def _group_interest_by_month(self, transactions: List[Transaction], _tax_year: str = None) -> Dict[str, Dict[str, Decimal]]:
        """
        Group interest transactions by month and loan account.

        Returns: {
            "2024-04": {"Loan 91-01": Decimal("500.00"), "Loan 91-03": Decimal("300.00")},
            ...
        }
        """
        monthly_data = defaultdict(lambda: defaultdict(Decimal))
        loan_pattern = re.compile(r'\b(\d{2}-\d{2})\b')

        # Categories that indicate interest - 'interest' is the primary code used by API
        interest_categories = {'interest', 'interest_expense', 'mortgage_interest', 'loan_interest'}

        for txn in transactions:
            if txn.category_code not in interest_categories or not txn.transaction_date:
                continue

            month_key = txn.transaction_date.strftime("%Y-%m")

            # Try to extract loan account
            loan_name = "Loan Account 1"  # Default
            if txn.description:
                matches = loan_pattern.findall(txn.description)
                if matches:
                    loan_name = f"Loan {matches[0]}"

            monthly_data[month_key][loan_name] += abs(txn.amount or Decimal(0))

        return dict(monthly_data)

    def _group_other_by_month(self, transactions: List[Transaction], _tax_year: str = None) -> Dict[str, Dict[str, Decimal]]:
        """
        Group rates, insurance, bank_fees transactions by month.

        Returns: {
            "2024-04": {"rates": Decimal("500.00"), "insurance": Decimal("200.00"), "bank_fees": Decimal("10.00")},
            ...
        }
        """
        monthly_data = defaultdict(lambda: defaultdict(Decimal))

        # Map various category codes to standard keys
        category_mapping = {
            'rates': 'rates',
            'council_rates': 'rates',
            'local_rates': 'rates',
            'insurance': 'insurance',
            'landlord_insurance': 'insurance',
            'property_insurance': 'insurance',
            'bank_fees': 'bank_fees',
            'bank_charges': 'bank_fees',
        }

        for txn in transactions:
            if not txn.transaction_date or not txn.category_code:
                continue

            standard_key = category_mapping.get(txn.category_code)
            if not standard_key:
                continue

            month_key = txn.transaction_date.strftime("%Y-%m")
            monthly_data[month_key][standard_key] += abs(txn.amount or Decimal(0))

        return dict(monthly_data)

    def _group_pm_by_month(self, transactions: List[Transaction], _tax_year: str = None) -> Dict[str, Dict[str, Decimal]]:
        """
        Group Property Manager transactions by month.

        Returns: {
            "2024-04": {"rental_income": Decimal("2000.00"), "agent_fees": Decimal("200.00"), "repairs_maintenance": Decimal("100.00")},
            ...
        }
        """
        monthly_data = defaultdict(lambda: defaultdict(Decimal))

        # Income categories
        income_cats = {'rental_income', 'rent', 'gross_rent'}
        # Agent fee categories
        agent_cats = {'agent_fees', 'property_management_fees', 'property_management', 'pm_fees', 'letting_fee'}
        # Repairs categories
        repairs_cats = {'repairs_maintenance', 'repairs', 'maintenance'}

        for txn in transactions:
            if not txn.transaction_date:
                continue

            # Check if from PM document
            doc_type = txn.document.document_type if txn.document else None
            if doc_type != "property_manager_statement":
                continue

            month_key = txn.transaction_date.strftime("%Y-%m")

            if txn.category_code in income_cats:
                monthly_data[month_key]["rental_income"] += abs(txn.amount or Decimal(0))
            elif txn.category_code in agent_cats:
                monthly_data[month_key]["agent_fees"] += abs(txn.amount or Decimal(0))
            elif txn.category_code in repairs_cats:
                monthly_data[month_key]["repairs_maintenance"] += abs(txn.amount or Decimal(0))

        return dict(monthly_data)

    def _get_repairs_items(self, transactions: List[Transaction]) -> List[Dict[str, Any]]:
        """Get individual repairs and maintenance items for detail section."""
        items = []
        repairs_cats = {'repairs_maintenance', 'repairs', 'maintenance'}

        for txn in transactions:
            if txn.category_code in repairs_cats and txn.amount:
                items.append({
                    "description": txn.description or txn.other_party or "Repair",
                    "amount": txn.amount,
                    "date": txn.transaction_date,
                })

        # Sort by amount descending (show largest items first)
        items.sort(key=lambda x: abs(x["amount"]), reverse=True)
        return items

    def _get_capital_items(self, transactions: List[Transaction]) -> List[Dict[str, Any]]:
        """Get capital expense items for detail section."""
        items = []
        capital_categories = {"capital_expense", "capital_purchase", "assets_over_500", "capital", "fixed_asset"}

        for txn in transactions:
            if txn.category_code in capital_categories and txn.amount:
                items.append({
                    "description": txn.description or txn.other_party or "Capital Item",
                    "amount": txn.amount,
                    "date": txn.transaction_date,
                })

        # Sort by amount descending
        items.sort(key=lambda x: abs(x["amount"]), reverse=True)
        return items

    # =========================================================================
    # TRANSACTION DETAIL SHEET METHODS
    # =========================================================================

    def _get_display_category(self, txn: Transaction) -> str:
        """
        Convert transaction category code to display name for workbook CODE column.

        Examples:
        - "rental_income" → "Rental Income"
        - "interest" with loan 91-01 → "Interest on Loan 91-01"
        - "rates" → "Rates"
        - "agent_fees" → "Property Management Fees"
        """
        if not txn.category_code:
            return "Uncategorized"

        code = txn.category_code.lower()

        # Special handling for interest - include loan account
        if "interest" in code and "principal" not in code:
            loan_account = self._extract_loan_account_from_txn(txn)
            if loan_account:
                return f"Interest on Loan {loan_account}"
            return "Interest Expense"

        # Special handling for principal
        if "principal" in code:
            loan_account = self._extract_loan_account_from_txn(txn)
            if loan_account:
                return f"Principal paid on {loan_account}"
            return "Principal Repayment"

        # Standard category mappings
        display_names = {
            "rental_income": "Rental Income",
            "rent_received": "Rental Income",
            "rent": "Rental Income",
            "water_rates_recovered": "Water Rates Recovered",
            "agent_fees": "Property Management Fees",
            "property_management_fees": "Property Management Fees",
            "property_management": "Property Management Fees",
            "rates": "Rates",
            "council_rates": "Rates",
            "water_rates": "Water Rates",
            "insurance": "Insurance",
            "landlord_insurance": "Landlord Insurance",
            "body_corporate": "Body Corporate",
            "body_corp": "Body Corporate",
            "repairs_maintenance": "Repairs & Maintenance",
            "repairs": "Repairs & Maintenance",
            "maintenance": "Repairs & Maintenance",
            "bank_fees": "Bank Fees",
            "accounting_fees": "Consulting & Accounting",
            "consulting_accounting": "Consulting & Accounting",
            "depreciation": "Depreciation",
            "advertising": "Advertising",
            "cleaning": "Cleaning",
            "legal_fees": "Legal Expenses",
            "legal_expenses": "Legal Expenses",
            "legal": "Legal Expenses",
            "due_diligence": "Due Diligence",
            "transfer": "Transfer",
            "bond": "Bond/Deposit",
            "funds_introduced": "Funds Introduced",
            "personal": "Personal Expense",
            "other_income": "Other Income",
            "bank_contribution": "Bank Contribution",
            "insurance_payout": "Insurance Payout",
            # Home office categories
            "ho_internet": "H/O Expenses - Internet",
            "ho_power": "H/O Expenses - Power",
            "ho_rates": "H/O Expenses - Rates",
            "ho_telephone": "H/O Expenses - Telephone",
            "ho_mobile": "H/O Expenses - Mobile",
        }

        return display_names.get(code, code.replace("_", " ").title())

    def _extract_loan_account_from_txn(self, txn: Transaction) -> Optional[str]:
        """Extract loan account number from transaction description/memo."""
        text = f"{txn.description or ''} {getattr(txn, 'memo', '') or ''}".lower()

        # Common patterns: "91-01", "91 001", "account 91-01"
        patterns = [
            re.compile(r'(\d{2}[-\s]?\d{2,3})'),  # 91-01, 91 001
            re.compile(r'account\s*#?\s*(\d+)', re.IGNORECASE),  # account 12345
        ]

        for pattern in patterns:
            match = pattern.search(text)
            if match:
                return match.group(1).replace(' ', '-')

        return None

    def _is_bank_transaction(self, txn: Transaction) -> bool:
        """Check if transaction is from bank statement."""
        # Check document type
        if txn.document:
            doc_type = (txn.document.document_type or '').lower()
            if 'bank' in doc_type or doc_type in ['bank_statement', 'bank_export']:
                return True

        # Check source document type attribute if exists
        source = getattr(txn, 'source_document_type', '') or ''
        source = source.lower()
        return 'bank' in source or source in ['bank_statement', 'bank_export']

    def _is_loan_transaction(self, txn: Transaction) -> bool:
        """Check if transaction is loan-related (interest or principal)."""
        cat = (txn.category_code or '').lower()
        return 'interest' in cat or 'principal' in cat

    def _is_pm_transaction(self, txn: Transaction) -> bool:
        """Check if transaction is from property manager."""
        # Check document type
        if txn.document:
            doc_type = (txn.document.document_type or '').lower()
            if 'pm' in doc_type or 'property_manager' in doc_type:
                return True

        # Check source document type attribute if exists
        source = getattr(txn, 'source_document_type', '') or ''
        source = source.lower()
        return 'pm' in source or 'property_manager' in source

    def _get_property_short_name(self, address: str) -> str:
        """Extract short property name from address for sheet naming.

        Excel sheet names cannot contain: / \\ ? * [ ] :
        Max length is 31 characters.
        """
        if not address:
            return "Property"
        # Take first part before comma
        short = address.split(',')[0].strip()
        # Remove invalid Excel sheet name characters
        invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            short = short.replace(char, '-')
        # Limit length (Excel max is 31, leaving room for suffix)
        if len(short) > 20:
            short = short[:20]
        return short

    def _get_fy_dates(self, tax_year: str) -> Tuple[datetime, datetime]:
        """Get financial year start and end dates from tax year string."""
        # Extract year from tax_year (e.g., "FY25" -> 2025)
        year = int("20" + tax_year[-2:])
        fy_start = datetime(year - 1, 4, 1)  # 1 April of previous year
        fy_end = datetime(year, 3, 31)  # 31 March of current year
        return fy_start, fy_end

    def _build_bank_statement_sheet(self, ws: Worksheet, transactions: List[Transaction],
                                    property_address: str, tax_year: str):
        """Build bank statement transaction register showing all bank transactions with category codes."""
        fy_start, fy_end = self._get_fy_dates(tax_year)

        # Header rows (matching original template format)
        ws["A1"] = f"Created date / time : {datetime.now().strftime('%d %B %Y / %H:%M:%S')}"
        ws["A2"] = f"Property: {property_address}"
        ws["A3"] = f"From date {fy_start.strftime('%Y%m%d')}"
        ws["A4"] = f"To date {fy_end.strftime('%Y%m%d')}"

        # Column headers (Row 6)
        headers = ["Date", "Unique Id", "Tran Type", "Cheque Number", "Payee", "Memo", "CODE", "Amount"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=6, column=col, value=header)
            ws.cell(row=6, column=col).font = Font(bold=True)

        # Column widths
        widths = [12, 12, 10, 14, 40, 40, 30, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Filter for bank statement transactions
        bank_txns = [t for t in transactions if self._is_bank_transaction(t)]
        bank_txns.sort(key=lambda t: t.transaction_date or datetime.min)

        # Data rows
        for row_num, txn in enumerate(bank_txns, 7):
            ws.cell(row=row_num, column=1, value=txn.transaction_date)
            ws.cell(row=row_num, column=1).number_format = 'DD/MM/YYYY'
            ws.cell(row=row_num, column=2, value=str(txn.id)[:8] if txn.id else '')
            ws.cell(row=row_num, column=3, value=getattr(txn, 'transaction_type', '') or '')
            ws.cell(row=row_num, column=4, value=getattr(txn, 'cheque_number', '') or '')
            ws.cell(row=row_num, column=5, value=txn.other_party or '')
            ws.cell(row=row_num, column=6, value=txn.description or '')
            ws.cell(row=row_num, column=7, value=self._get_display_category(txn))  # THE CODE
            ws.cell(row=row_num, column=8, value=float(txn.amount) if txn.amount else 0)
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'

        # Add totals row
        total_row = 7 + len(bank_txns)
        ws.cell(row=total_row, column=7, value="TOTAL")
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=f"=SUM(H7:H{total_row-1})")
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=8).number_format = '#,##0.00'

        logger.info(f"Bank Statement sheet: {len(bank_txns)} transactions")

    def _build_loan_statement_sheet(self, ws: Worksheet, transactions: List[Transaction],
                                    property_address: str, tax_year: str):
        """Build loan statement showing interest and principal transactions with breakdown."""
        fy_start, fy_end = self._get_fy_dates(tax_year)

        # Header
        ws["A1"] = f"Created date / time : {datetime.now().strftime('%d %B %Y / %H:%M:%S')}"
        ws["A2"] = f"Property: {property_address}"
        ws["A3"] = f"From date {fy_start.strftime('%Y%m%d')}"
        ws["A4"] = f"To date {fy_end.strftime('%Y%m%d')}"

        # Column headers
        headers = ["Date", "Unique Id", "Tran Type", "Cheque Number", "Payee", "Memo", "CODE", "Amount"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=6, column=col, value=header)
            ws.cell(row=6, column=col).font = Font(bold=True)

        # Column widths
        widths = [12, 12, 12, 14, 25, 35, 30, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Filter for loan transactions (interest and principal)
        loan_txns = [t for t in transactions if self._is_loan_transaction(t)]

        # Sort by date, then by type (interest first)
        def sort_key(t):
            date = t.transaction_date or datetime.min
            is_principal = 'principal' in (t.category_code or '').lower()
            return (date, is_principal)

        loan_txns.sort(key=sort_key)

        # Data rows
        for row_num, txn in enumerate(loan_txns, 7):
            ws.cell(row=row_num, column=1, value=txn.transaction_date)
            ws.cell(row=row_num, column=1).number_format = 'DD/MM/YYYY'
            ws.cell(row=row_num, column=2, value=str(txn.id)[:8] if txn.id else '')

            # Determine transaction type
            if 'interest' in (txn.category_code or '').lower() and 'principal' not in (txn.category_code or '').lower():
                tran_type = "LOAN INT"
            elif 'principal' in (txn.category_code or '').lower():
                tran_type = "LOAN PRIN"
            else:
                tran_type = getattr(txn, 'transaction_type', '') or ''

            ws.cell(row=row_num, column=3, value=tran_type)
            ws.cell(row=row_num, column=4, value='')
            ws.cell(row=row_num, column=5, value=txn.other_party or 'LOAN')
            ws.cell(row=row_num, column=6, value=txn.description or '')
            ws.cell(row=row_num, column=7, value=self._get_display_category(txn))
            ws.cell(row=row_num, column=8, value=float(txn.amount) if txn.amount else 0)
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'

        # Add summary section
        summary_start = 7 + len(loan_txns) + 2

        # Interest total
        interest_txns = [t for t in loan_txns if 'interest' in (t.category_code or '').lower()
                         and 'principal' not in (t.category_code or '').lower()]
        interest_total = sum(abs(t.amount or 0) for t in interest_txns)

        # Principal total
        principal_txns = [t for t in loan_txns if 'principal' in (t.category_code or '').lower()]
        principal_total = sum(abs(t.amount or 0) for t in principal_txns)

        ws.cell(row=summary_start, column=6, value="Interest Total:")
        ws.cell(row=summary_start, column=6).font = Font(bold=True)
        ws.cell(row=summary_start, column=8, value=float(interest_total))
        ws.cell(row=summary_start, column=8).font = Font(bold=True)
        ws.cell(row=summary_start, column=8).number_format = '#,##0.00'

        ws.cell(row=summary_start + 1, column=6, value="Principal Total:")
        ws.cell(row=summary_start + 1, column=6).font = Font(bold=True)
        ws.cell(row=summary_start + 1, column=8, value=float(principal_total))
        ws.cell(row=summary_start + 1, column=8).font = Font(bold=True)
        ws.cell(row=summary_start + 1, column=8).number_format = '#,##0.00'

        logger.info(f"Loan Statement sheet: {len(loan_txns)} transactions (interest: {len(interest_txns)}, principal: {len(principal_txns)})")

    def _build_pm_statement_sheet(self, ws: Worksheet, transactions: List[Transaction],
                                  property_address: str, tax_year: str):
        """Build property manager statement showing PM transactions with categories."""
        fy_start, fy_end = self._get_fy_dates(tax_year)

        # Header
        ws["A1"] = f"Created date / time : {datetime.now().strftime('%d %B %Y / %H:%M:%S')}"
        ws["A2"] = f"Property: {property_address}"
        ws["A3"] = f"From date {fy_start.strftime('%Y%m%d')}"
        ws["A4"] = f"To date {fy_end.strftime('%Y%m%d')}"

        # Column headers
        headers = ["Date", "Description", "CODE", "Amount"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=6, column=col, value=header)
            ws.cell(row=6, column=col).font = Font(bold=True)

        # Column widths
        widths = [12, 50, 30, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Filter for PM transactions
        pm_txns = [t for t in transactions if self._is_pm_transaction(t)]
        pm_txns.sort(key=lambda t: t.transaction_date or datetime.min)

        # Data rows
        for row_num, txn in enumerate(pm_txns, 7):
            ws.cell(row=row_num, column=1, value=txn.transaction_date)
            ws.cell(row=row_num, column=1).number_format = 'DD/MM/YYYY'
            ws.cell(row=row_num, column=2, value=txn.description or txn.other_party or '')
            ws.cell(row=row_num, column=3, value=self._get_display_category(txn))
            ws.cell(row=row_num, column=4, value=float(txn.amount) if txn.amount else 0)
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'

        # Add totals row
        total_row = 7 + len(pm_txns)
        ws.cell(row=total_row, column=3, value="TOTAL")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=f"=SUM(D7:D{total_row-1})")
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        ws.cell(row=total_row, column=4).number_format = '#,##0.00'

        logger.info(f"PM Statement sheet: {len(pm_txns)} transactions")

    def _build_fy_summary_sheet(self, ws: Worksheet, context: Dict[str, Any]):
        """Build FY summary pivot showing totals by category for verification."""
        tax_return = context["tax_return"]
        transactions = context["transactions"]

        ws["A1"] = f"FY{tax_return.tax_year} Summary - {tax_return.property_address}"
        ws["A1"].font = Font(bold=True, size=14)

        ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M:%S')}"

        # Group transactions by category
        category_totals: Dict[str, Decimal] = defaultdict(Decimal)
        category_counts: Dict[str, int] = defaultdict(int)

        for txn in transactions:
            cat = txn.category_code or "uncategorized"
            category_totals[cat] += txn.amount or Decimal(0)
            category_counts[cat] += 1

        # Headers
        ws["A4"] = "Category"
        ws["B4"] = "Display Name"
        ws["C4"] = "Count"
        ws["D4"] = "Total Amount"
        for col in range(1, 5):
            ws.cell(row=4, column=col).font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15

        # Income categories
        income_cats = ['rental_income', 'rent', 'water_rates_recovered', 'bank_contribution',
                       'insurance_payout', 'other_income']

        row = 6

        # Income section
        ws[f"A{row}"] = "INCOME"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        income_total = Decimal(0)
        for cat in income_cats:
            if cat in category_totals and category_totals[cat] > 0:
                ws[f"A{row}"] = cat
                ws[f"B{row}"] = self._get_display_category_name(cat)
                ws[f"C{row}"] = category_counts[cat]
                ws[f"D{row}"] = float(abs(category_totals[cat]))
                ws[f"D{row}"].number_format = '#,##0.00'
                income_total += abs(category_totals[cat])
                row += 1

        ws[f"A{row}"] = "Total Income"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"D{row}"] = float(income_total)
        ws[f"D{row}"].font = Font(bold=True)
        ws[f"D{row}"].number_format = '#,##0.00'
        row += 2

        # Expenses section
        ws[f"A{row}"] = "EXPENSES"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        expense_total = Decimal(0)
        excluded_cats = set(income_cats) | {'uncategorized', 'transfer', 'bond', 'funds_introduced', 'personal', 'unknown'}

        for cat, total in sorted(category_totals.items()):
            if cat not in excluded_cats and total < 0:
                ws[f"A{row}"] = cat
                ws[f"B{row}"] = self._get_display_category_name(cat)
                ws[f"C{row}"] = category_counts[cat]
                ws[f"D{row}"] = float(abs(total))
                ws[f"D{row}"].number_format = '#,##0.00'
                expense_total += abs(total)
                row += 1

        ws[f"A{row}"] = "Total Expenses"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"D{row}"] = float(expense_total)
        ws[f"D{row}"].font = Font(bold=True)
        ws[f"D{row}"].number_format = '#,##0.00'
        row += 2

        # Excluded items section
        ws[f"A{row}"] = "EXCLUDED (Non-deductible)"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        for cat in ['transfer', 'bond', 'funds_introduced', 'personal', 'principal_repayment', 'unknown', 'uncategorized']:
            if cat in category_totals:
                ws[f"A{row}"] = cat
                ws[f"B{row}"] = self._get_display_category_name(cat)
                ws[f"C{row}"] = category_counts[cat]
                ws[f"D{row}"] = float(category_totals[cat])
                ws[f"D{row}"].number_format = '#,##0.00'
                row += 1

        row += 2
        # Net summary
        ws[f"A{row}"] = "NET RENTAL INCOME"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"D{row}"] = float(income_total - expense_total)
        ws[f"D{row}"].font = Font(bold=True, size=12)
        ws[f"D{row}"].number_format = '#,##0.00'

        logger.info(f"FY Summary sheet: {len(category_totals)} categories, Income: ${income_total:,.2f}, Expenses: ${expense_total:,.2f}")

    def _get_display_category_name(self, cat: str) -> str:
        """Get display name for a category code (for FY summary)."""
        display_names = {
            "rental_income": "Rental Income",
            "rent": "Rental Income",
            "water_rates_recovered": "Water Rates Recovered",
            "bank_contribution": "Bank Contribution",
            "insurance_payout": "Insurance Payout",
            "other_income": "Other Income",
            "interest": "Interest Expense",
            "interest_expense": "Interest Expense",
            "principal_repayment": "Principal Repayment (non-deductible)",
            "rates": "Council Rates",
            "water_rates": "Water Rates",
            "agent_fees": "Property Management Fees",
            "property_management_fees": "Property Management Fees",
            "body_corporate": "Body Corporate",
            "insurance": "Insurance",
            "repairs_maintenance": "Repairs & Maintenance",
            "bank_fees": "Bank Fees",
            "depreciation": "Depreciation",
            "legal_fees": "Legal Expenses",
            "advertising": "Advertising",
            "transfer": "Transfers (excluded)",
            "bond": "Bond/Deposit (excluded)",
            "funds_introduced": "Funds Introduced (excluded)",
            "personal": "Personal (excluded)",
            "unknown": "Unknown/Uncategorized",
            "uncategorized": "Uncategorized",
        }
        return display_names.get(cat, cat.replace("_", " ").title())

    # =========================================================================
    # DATA LOADING METHODS
    # =========================================================================

    async def _load_tax_return(self, db: AsyncSession, tax_return_id: UUID) -> TaxReturn:
        """Load tax return with client."""
        result = await db.execute(
            select(TaxReturn)
            .options(selectinload(TaxReturn.client))
            .where(TaxReturn.id == tax_return_id)
        )
        tax_return = result.scalar_one_or_none()
        if not tax_return:
            raise ValueError(f"Tax return not found: {tax_return_id}")
        return tax_return

    async def _load_transactions_with_documents(self, db: AsyncSession, tax_return_id: UUID) -> List[Transaction]:
        """Load all transactions with their source documents."""
        result = await db.execute(
            select(Transaction)
            .options(selectinload(Transaction.document))
            .where(Transaction.tax_return_id == tax_return_id)
            .order_by(Transaction.transaction_date)
        )
        return list(result.scalars().all())

    async def _load_summaries(self, db: AsyncSession, tax_return_id: UUID) -> List[TransactionSummary]:
        """Load transaction summaries with category mappings."""
        result = await db.execute(
            select(TransactionSummary)
            .options(selectinload(TransactionSummary.category_mapping))
            .where(TransactionSummary.tax_return_id == tax_return_id)
        )
        return list(result.scalars().all())

    async def _load_pl_mappings(self, db: AsyncSession) -> List[PLRowMapping]:
        """Load P&L mappings."""
        result = await db.execute(
            select(PLRowMapping).order_by(PLRowMapping.sort_order)
        )
        return list(result.scalars().all())

    async def _load_workings(self, db: AsyncSession, tax_return_id: UUID) -> Optional[TaxReturnWorkingsData]:
        """Load AI Brain workings from database."""
        result = await db.execute(
            select(TaxReturnWorkings)
            .where(TaxReturnWorkings.tax_return_id == tax_return_id)
            .order_by(TaxReturnWorkings.version.desc())
            .limit(1)
        )
        db_workings = result.scalar_one_or_none()

        if not db_workings:
            logger.warning(f"No AI Brain workings found for tax return {tax_return_id}")
            return None

        # Check if we have the workings data (income_workings and expense_workings are separate JSONB columns)
        if not db_workings.income_workings and not db_workings.expense_workings:
            logger.warning(f"AI Brain workings record exists but no income/expense data for tax return {tax_return_id}")
            return None

        try:
            # Build IncomeWorkings from JSONB
            income_data = db_workings.income_workings or {}
            income_workings = IncomeWorkings.model_validate(income_data)

            # Build ExpenseWorkings from JSONB
            expense_data = db_workings.expense_workings or {}
            expense_workings = ExpenseWorkings.model_validate(expense_data)

            # Build WorkingsSummary from individual columns
            summary = WorkingsSummary(
                total_income=Decimal(str(db_workings.total_income or 0)),
                total_expenses=Decimal(str(db_workings.total_expenses or 0)),
                total_deductions=Decimal(str(db_workings.total_deductions or 0)),
                interest_gross=Decimal(str(db_workings.interest_gross or 0)),
                interest_deductible_percentage=float(db_workings.interest_deductible_percentage or 80.0),
                interest_deductible_amount=Decimal(str(db_workings.interest_deductible_amount or 0)),
                net_rental_income=Decimal(str(db_workings.net_rental_income or 0)),
            )

            # Load tax return for property info
            tax_return = await self._load_tax_return(db, tax_return_id)

            # Build TaxReturnWorkingsData
            workings = TaxReturnWorkingsData(
                tax_return_id=tax_return_id,
                property_address=tax_return.property_address,
                tax_year=tax_return.tax_year,
                property_type=tax_return.property_type.value,
                summary=summary,
                income=income_workings,
                expenses=expense_workings,
                processing_notes=db_workings.processing_notes or [],
                audit_trail=db_workings.audit_trail or [],
            )

            logger.info(f"Loaded AI Brain workings v{db_workings.version} for tax return {tax_return_id}")
            logger.info(f"  - Total Income: ${summary.total_income:,.2f}")
            logger.info(f"  - Total Deductions: ${summary.total_deductions:,.2f}")
            logger.info(f"  - Interest: gross=${summary.interest_gross:,.2f}, deductible=${summary.interest_deductible_amount:,.2f} ({summary.interest_deductible_percentage:.0f}%)")
            logger.info(f"  - Net Rental Income: ${summary.net_rental_income:,.2f}")
            return workings
        except Exception as e:
            logger.error(f"Failed to parse workings data: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None


# Singleton
_workbook_generator: Optional[WorkbookGenerator] = None


def get_workbook_generator() -> WorkbookGenerator:
    """Get singleton workbook generator."""
    global _workbook_generator
    if _workbook_generator is None:
        _workbook_generator = WorkbookGenerator()
    return _workbook_generator
