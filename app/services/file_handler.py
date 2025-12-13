"""File handling service for processing uploaded documents."""

import csv
import hashlib
import io
import logging
import uuid
from pathlib import Path
from typing import Tuple

import fitz  # PyMuPDF
import openpyxl
import pypdf
import xlrd
from fastapi import HTTPException, UploadFile
from PIL import Image

from app.config import settings
from app.schemas.documents import ProcessedFile

logger = logging.getLogger(__name__)


class FileHandler:
    """Handle file uploads and processing."""

    def __init__(self):
        """Initialize file handler."""
        self.upload_dir = settings.UPLOAD_DIR
        self.max_file_size = settings.max_file_size_bytes

    async def save_upload(self, file: UploadFile, tax_return_id: str) -> Tuple[str, str, int, str]:
        """
        Save uploaded file to disk.

        Args:
            file: Uploaded file
            tax_return_id: Tax return ID for directory organization

        Returns:
            Tuple of (stored_filename, file_path, file_size, content_hash)
        """
        # Validate file
        await self._validate_file(file)

        # Create directory for tax return
        tax_return_dir = self.upload_dir / tax_return_id
        tax_return_dir.mkdir(parents=True, exist_ok=True)

        # Generate unique filename
        stored_filename = f"{uuid.uuid4()}_{file.filename}"
        file_path = tax_return_dir / stored_filename

        # Save file and compute hash
        content = await file.read()
        content_hash = self._compute_hash(content)
        file_path.write_bytes(content)

        return stored_filename, str(file_path), len(content), content_hash

    def _compute_hash(self, content: bytes) -> str:
        """Compute SHA-256 hash of file content."""
        return hashlib.sha256(content).hexdigest()

    async def process_file(self, file_path: str, original_filename: str) -> ProcessedFile:
        """
        Process a saved file and extract content.

        Args:
            file_path: Path to saved file
            original_filename: Original filename for type detection

        Returns:
            ProcessedFile object with extracted content
        """
        file_path_obj = Path(file_path)
        file_ext = Path(original_filename).suffix.lower()

        if file_ext == ".pdf":
            return await self._process_pdf(file_path_obj)
        elif file_ext in [".png", ".jpg", ".jpeg"]:
            return await self._process_image(file_path_obj)
        elif file_ext == ".xlsx":
            return await self._process_excel(file_path_obj)
        elif file_ext == ".xls":
            return await self._process_xls(file_path_obj)
        elif file_ext == ".csv":
            return await self._process_csv(file_path_obj)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_ext}")

    async def _validate_file(self, file: UploadFile) -> None:
        """Validate uploaded file."""
        # Check file extension
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in settings.ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"File type {file_ext} not allowed. Allowed types: {settings.ALLOWED_EXTENSIONS}",
            )

        # Check MIME type
        if file.content_type not in settings.ALLOWED_MIME_TYPES:
            raise HTTPException(
                status_code=400, detail=f"MIME type {file.content_type} not allowed"
            )

        # Check file size
        file.file.seek(0, 2)  # Move to end
        file_size = file.file.tell()
        file.file.seek(0)  # Reset

        if file_size > self.max_file_size:
            raise HTTPException(
                status_code=400,
                detail=f"File size ({file_size / 1024 / 1024:.1f}MB) exceeds maximum ({settings.MAX_FILE_SIZE_MB}MB)",
            )

    async def _process_pdf(self, file_path: Path) -> ProcessedFile:
        """Process PDF file."""
        content = file_path.read_bytes()

        try:
            # Try text extraction first
            pdf_reader = pypdf.PdfReader(io.BytesIO(content))
            text_content = []
            page_count = len(pdf_reader.pages)

            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_content.append(page_text)

            combined_text = "\n".join(text_content)

            # Check if PDF is scanned (minimal/garbled text)
            words = combined_text.split()
            # Scanned if: very little text, OR text looks garbled (low ratio of real words)
            is_scanned = (
                len(combined_text.strip()) < 200  # Very little text
                or len(words) < 20  # Very few words
                or sum(1 for w in words if len(w) > 15) / max(len(words), 1)
                > 0.3  # Many long "words" = garbled OCR
            )

            if is_scanned:
                logger.info(f"PDF appears to be scanned or has poor text extraction: {file_path}")
                return await self._convert_pdf_to_images(file_path, content, page_count)

            # Digital PDF with good text extraction
            return ProcessedFile(
                file_path=str(file_path),
                file_type="digital_pdf",
                text_content=combined_text,
                image_paths=None,
                page_count=page_count,
            )

        except Exception as e:
            logger.error(f"Error processing PDF: {e}")
            # If text extraction fails, try converting to images
            return await self._convert_pdf_to_images(file_path, content, 0)

    async def _convert_pdf_to_images(
        self, file_path: Path, content: bytes, page_count: int
    ) -> ProcessedFile:
        """Convert PDF pages to images using PyMuPDF."""
        try:
            # Use PyMuPDF (fitz) - no external dependencies like poppler required
            pdf_document = fitz.open(stream=content, filetype="pdf")
            image_paths = []

            # Convert each page to image (300 DPI equivalent)
            zoom = 300 / 72  # 72 is default PDF DPI, we want 300 DPI
            matrix = fitz.Matrix(zoom, zoom)

            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                pix = page.get_pixmap(matrix=matrix)

                image_path = file_path.parent / f"{file_path.stem}_page_{page_num + 1}.png"
                pix.save(str(image_path))
                image_paths.append(str(image_path))

            pdf_document.close()

            return ProcessedFile(
                file_path=str(file_path),
                file_type="scanned_pdf",
                text_content=None,
                image_paths=image_paths,
                page_count=page_count or len(image_paths),
            )

        except Exception as e:
            logger.error(f"Error converting PDF to images: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to process PDF file: {str(e)}")

    async def _process_image(self, file_path: Path) -> ProcessedFile:
        """Process image file."""
        try:
            # Verify it's a valid image
            with Image.open(file_path) as img:
                width, height = img.size

            return ProcessedFile(
                file_path=str(file_path),
                file_type="image",
                text_content=None,
                image_paths=[str(file_path)],
                page_count=1,
            )

        except Exception as e:
            logger.error(f"Error processing image: {e}")
            raise HTTPException(status_code=400, detail=f"Invalid image file: {str(e)}")

    async def _process_excel(self, file_path: Path) -> ProcessedFile:
        """Process Excel file."""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_content = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"Sheet: {sheet_name}\n")

                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_data):  # Skip empty rows
                        text_content.append("\t".join(row_data))

                text_content.append("\n")  # Add spacing between sheets

            return ProcessedFile(
                file_path=str(file_path),
                file_type="spreadsheet",
                text_content="\n".join(text_content),
                image_paths=None,
                page_count=len(workbook.sheetnames),
            )

        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    async def _process_xls(self, file_path: Path) -> ProcessedFile:
        """Process old .xls Excel file."""
        try:
            workbook = xlrd.open_workbook(file_path)
            text_content = []

            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                text_content.append(f"Sheet: {sheet_name}\n")

                for row_idx in range(min(sheet.nrows, 500)):  # Limit rows
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_data.append(str(cell_value) if cell_value else "")
                    if any(row_data):
                        text_content.append("\t".join(row_data))

                text_content.append("\n")

            return ProcessedFile(
                file_path=str(file_path),
                file_type="spreadsheet",
                text_content="\n".join(text_content),
                image_paths=None,
                page_count=len(workbook.sheet_names()),
            )

        except Exception as e:
            logger.error(f"Error processing XLS file: {e}")
            raise HTTPException(status_code=400, detail=f"Invalid XLS file: {str(e)}")

    async def _process_csv(self, file_path: Path) -> ProcessedFile:
        """Process CSV file."""
        try:
            content = file_path.read_text(encoding="utf-8", errors="ignore")
            rows = []

            reader = csv.reader(io.StringIO(content))
            for row in reader:
                if row:  # Skip empty rows
                    rows.append("\t".join(row))

            return ProcessedFile(
                file_path=str(file_path),
                file_type="spreadsheet",
                text_content="\n".join(rows),
                image_paths=None,
                page_count=1,
            )

        except Exception as e:
            logger.error(f"Error processing CSV file: {e}")
            raise HTTPException(status_code=400, detail=f"Invalid CSV file: {str(e)}")
