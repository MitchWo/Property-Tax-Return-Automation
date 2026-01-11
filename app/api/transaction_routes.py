"""API routes for transaction management."""
import asyncio
import logging
from decimal import Decimal
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db, AsyncSessionLocal
from app.services.progress_tracker import create_tracker, get_tracker, remove_tracker
from app.models.db_models import (
    CategoryFeedback,
    PLRowMapping,
    TaxReturn,
    Transaction,
    TransactionPattern,
    TransactionSummary,
)
from app.schemas.transactions import (
    CategoryFeedbackCreate,
    CategoryFeedbackResponse,
    MatchingTransactionInfo,
    PLRowMappingResponse,
    TransactionBulkUpdate,
    TransactionListResponse,
    TransactionPatternResponse,
    TransactionResponse,
    TransactionSummaryResponse,
    TransactionUpdate,
    TransactionUpdateResponse,
)
from app.services.rag_categorization_integration import get_rag_integration
from app.services.transaction_categorizer import get_transaction_categorizer
from app.services.transaction_processor import TransactionProcessor
from app.services.phase2_feedback_learning.knowledge_store import knowledge_store
from app.schemas.transactions import ProcessingResult

logger = logging.getLogger(__name__)

# Create router
transaction_router = APIRouter(prefix="/api/transactions", tags=["transactions"])
transaction_web_router = APIRouter(tags=["transactions-web"])


# =============================================================================
# API ROUTES
# =============================================================================

@transaction_router.get("/return/{tax_return_id}", response_model=TransactionListResponse)
async def get_transactions_for_return(
    tax_return_id: UUID,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    category_code: Optional[str] = None,
    needs_review: Optional[bool] = None,
    transaction_type: Optional[str] = None,
    min_amount: Optional[float] = None,
    max_amount: Optional[float] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """Get transactions for a tax return with filtering and pagination."""
    # Build query
    query = select(Transaction).where(Transaction.tax_return_id == tax_return_id)

    # Apply filters
    if category_code:
        query = query.where(Transaction.category_code == category_code)
    if needs_review is not None:
        query = query.where(Transaction.needs_review == needs_review)
    if transaction_type:
        query = query.where(Transaction.transaction_type == transaction_type)
    if min_amount is not None:
        query = query.where(func.abs(Transaction.amount) >= min_amount)
    if max_amount is not None:
        query = query.where(func.abs(Transaction.amount) <= max_amount)
    if search:
        query = query.where(Transaction.description.ilike(f"%{search}%"))

    # Get total count
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar()

    # Apply pagination
    offset = (page - 1) * page_size
    query = query.order_by(Transaction.transaction_date.desc()).offset(offset).limit(page_size)

    result = await db.execute(query)
    transactions = result.scalars().all()

    # Get P&L mappings for display names
    mappings_result = await db.execute(select(PLRowMapping))
    mappings = {m.category_code: m for m in mappings_result.scalars().all()}

    # Build response
    transaction_responses = []
    for txn in transactions:
        mapping = mappings.get(txn.category_code)
        transaction_responses.append(TransactionResponse(
            id=txn.id,
            tax_return_id=txn.tax_return_id,
            document_id=txn.document_id,
            transaction_date=txn.transaction_date,
            description=txn.description,
            other_party=txn.other_party,
            amount=txn.amount,
            balance=txn.balance,
            category_code=txn.category_code,
            category_display_name=mapping.display_name if mapping else None,
            transaction_type=txn.transaction_type,
            pl_row=mapping.pl_row if mapping else None,
            is_deductible=txn.is_deductible,
            deductible_percentage=txn.deductible_percentage,
            deductible_amount=txn.deductible_amount,
            gst_inclusive=txn.gst_inclusive,
            gst_amount=txn.gst_amount,
            confidence=txn.confidence,
            categorization_source=txn.categorization_source,
            needs_review=txn.needs_review,
            review_reason=txn.review_reason,
            manually_reviewed=txn.manually_reviewed,
            reviewed_by=txn.reviewed_by,
            reviewed_at=txn.reviewed_at,
            created_at=txn.created_at,
            updated_at=txn.updated_at
        ))

    # Calculate summary stats
    all_txns_query = select(Transaction).where(Transaction.tax_return_id == tax_return_id)
    all_result = await db.execute(all_txns_query)
    all_txns = all_result.scalars().all()

    total_income = sum(t.amount for t in all_txns if t.amount > 0)
    total_expenses = sum(abs(t.amount) for t in all_txns if t.amount < 0)
    needs_review_count = sum(1 for t in all_txns if t.needs_review)

    return TransactionListResponse(
        transactions=transaction_responses,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=(total + page_size - 1) // page_size,
        total_income=Decimal(str(total_income)),
        total_expenses=Decimal(str(total_expenses)),
        needs_review_count=needs_review_count
    )


@transaction_router.get("/{transaction_id}", response_model=TransactionResponse)
async def get_transaction(
    transaction_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """Get a single transaction by ID."""
    result = await db.execute(
        select(Transaction).where(Transaction.id == transaction_id)
    )
    txn = result.scalar_one_or_none()

    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # Get mapping
    mapping_result = await db.execute(
        select(PLRowMapping).where(PLRowMapping.category_code == txn.category_code)
    )
    mapping = mapping_result.scalar_one_or_none()

    return TransactionResponse(
        id=txn.id,
        tax_return_id=txn.tax_return_id,
        document_id=txn.document_id,
        transaction_date=txn.transaction_date,
        description=txn.description,
        other_party=txn.other_party,
        amount=txn.amount,
        balance=txn.balance,
        category_code=txn.category_code,
        category_display_name=mapping.display_name if mapping else None,
        transaction_type=txn.transaction_type,
        pl_row=mapping.pl_row if mapping else None,
        is_deductible=txn.is_deductible,
        deductible_percentage=txn.deductible_percentage,
        deductible_amount=txn.deductible_amount,
        gst_inclusive=txn.gst_inclusive,
        gst_amount=txn.gst_amount,
        confidence=txn.confidence,
        categorization_source=txn.categorization_source,
        needs_review=txn.needs_review,
        review_reason=txn.review_reason,
        manually_reviewed=txn.manually_reviewed,
        reviewed_by=txn.reviewed_by,
        reviewed_at=txn.reviewed_at,
        created_at=txn.created_at,
        updated_at=txn.updated_at
    )


@transaction_router.put("/{transaction_id}", response_model=TransactionUpdateResponse)
async def update_transaction(
    transaction_id: UUID,
    update: TransactionUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update a transaction (typically for category correction).

    Returns the updated transaction plus any matching transactions
    (same vendor + same amount) that could be bulk-updated.
    """
    logger.info(f"[UPDATE] Received PUT for transaction {transaction_id}")
    logger.info(f"[UPDATE] Request data: category_code={update.category_code}")

    result = await db.execute(
        select(Transaction).where(Transaction.id == transaction_id)
    )
    txn = result.scalar_one_or_none()

    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # Store original values for matching
    original_category = txn.category_code
    logger.info(f"[UPDATE] Original category: {original_category} -> New: {update.category_code}")
    txn_other_party = txn.other_party
    txn_amount = txn.amount

    # If category is being changed, learn from the correction
    if update.category_code and update.category_code != txn.category_code:
        logger.info("[UPDATE] Category change detected, calling learn_from_correction")
        # Use the existing categorizer learning
        categorizer = get_transaction_categorizer()
        await categorizer.learn_from_correction(
            db=db,
            transaction_id=transaction_id,
            corrected_category=update.category_code,
            corrected_by=None,  # Would come from auth in production
            notes=update.notes,
            create_pattern=True
        )
        logger.info("[UPDATE] learn_from_correction completed")

        # === Also create RAG learning for semantic search ===
        try:
            rag_integration = get_rag_integration(db)
            learning = await rag_integration.learn_from_correction(
                transaction=txn,
                original_category=original_category,
                corrected_category=update.category_code,
                corrected_by="user",  # Would come from auth in production
                notes=update.notes
            )

            if learning:
                logger.info(f"Auto-created RAG learning from correction: '{txn.description}' → {update.category_code}")
        except Exception as e:
            logger.warning(f"Failed to create RAG learning from correction: {e}")
            # Don't fail the request - RAG learning is a bonus
    else:
        # Just update the fields
        if update.category_code is not None:
            txn.category_code = update.category_code
        if update.transaction_type is not None:
            txn.transaction_type = update.transaction_type
        if update.is_deductible is not None:
            txn.is_deductible = update.is_deductible
        if update.deductible_percentage is not None:
            txn.deductible_percentage = update.deductible_percentage
        if update.needs_review is not None:
            txn.needs_review = update.needs_review
        if update.review_reason is not None:
            txn.review_reason = update.review_reason

        await db.commit()
        await db.refresh(txn)

    # Get updated transaction
    updated_txn = await get_transaction(transaction_id, db)
    logger.info(f"[UPDATE] Returning updated transaction with category: {updated_txn.category_code}")

    # Find matching transactions (same vendor + same amount, different category)
    matching_transactions = []
    prompt_message = None

    if update.category_code and txn_other_party and txn_amount:
        # Find other transactions with same vendor and amount that have a different category
        matching_result = await db.execute(
            select(Transaction).where(
                Transaction.tax_return_id == txn.tax_return_id,
                Transaction.id != transaction_id,
                Transaction.other_party == txn_other_party,
                Transaction.amount == txn_amount,
                Transaction.category_code != update.category_code
            )
        )
        matching_txns = matching_result.scalars().all()

        if matching_txns:
            matching_transactions = [
                MatchingTransactionInfo(
                    id=m.id,
                    transaction_date=m.transaction_date,
                    description=m.description,
                    other_party=m.other_party,
                    amount=m.amount,
                    current_category=m.category_code
                )
                for m in matching_txns
            ]

            count = len(matching_transactions)
            prompt_message = f"Found {count} other transaction{'s' if count > 1 else ''} to '{txn_other_party}' for ${abs(txn_amount):.2f}. Would you like to apply the same category to all of them?"

    return TransactionUpdateResponse(
        transaction=updated_txn,
        matching_transactions=matching_transactions,
        matching_count=len(matching_transactions),
        prompt_message=prompt_message
    )


@transaction_router.post("/{transaction_id}/confirm")
async def confirm_transaction(
    transaction_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """
    Confirm a transaction's AI-suggested category is correct.

    This marks the transaction as reviewed without changing its category.
    Useful when the AI's suggestion is correct and user wants to approve it.
    """
    from datetime import datetime, timezone

    result = await db.execute(
        select(Transaction).where(Transaction.id == transaction_id)
    )
    txn = result.scalar_one_or_none()

    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # Mark as reviewed without changing category
    txn.needs_review = False
    txn.manually_reviewed = True
    txn.reviewed_at = datetime.now(timezone.utc)
    txn.review_reason = None  # Clear the review reason since it's been confirmed

    await db.commit()
    await db.refresh(txn)

    logger.info(f"Transaction {transaction_id} confirmed with category '{txn.category_code}'")

    return {
        "id": str(txn.id),
        "category_code": txn.category_code,
        "needs_review": txn.needs_review,
        "manually_reviewed": txn.manually_reviewed,
        "message": "Transaction confirmed successfully"
    }


@transaction_router.post("/bulk-confirm")
async def bulk_confirm_transactions(
    transaction_ids: List[UUID],
    db: AsyncSession = Depends(get_db)
):
    """
    Bulk confirm multiple transactions' AI-suggested categories.

    Marks all specified transactions as reviewed without changing their categories.
    """
    from datetime import datetime, timezone

    confirmed_count = 0

    for txn_id in transaction_ids:
        result = await db.execute(
            select(Transaction).where(Transaction.id == txn_id)
        )
        txn = result.scalar_one_or_none()

        if txn:
            txn.needs_review = False
            txn.manually_reviewed = True
            txn.reviewed_at = datetime.now(timezone.utc)
            txn.review_reason = None
            confirmed_count += 1

    await db.commit()

    logger.info(f"Bulk confirmed {confirmed_count} transactions")

    return {
        "confirmed_count": confirmed_count,
        "message": f"Confirmed {confirmed_count} transactions"
    }


@transaction_router.post("/bulk-update")
async def bulk_update_transactions(
    update: TransactionBulkUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Bulk update multiple transactions to the same category."""
    categorizer = get_transaction_categorizer()
    rag_integration = get_rag_integration(db)

    updated_count = 0
    pattern_id = None
    rag_learnings_created = 0

    for txn_id in update.transaction_ids:
        try:
            # Get the transaction for RAG learning
            result = await db.execute(
                select(Transaction).where(Transaction.id == txn_id)
            )
            txn = result.scalar_one_or_none()

            if not txn:
                logger.warning(f"Transaction {txn_id} not found")
                continue

            original_category = txn.category_code

            # Use existing categorizer learning
            result = await categorizer.learn_from_correction(
                db=db,
                transaction_id=txn_id,
                corrected_category=update.category_code,
                create_pattern=update.apply_to_similar
            )
            if result and not pattern_id:
                pattern_id = result

            # === Create RAG learning for first transaction only (to avoid duplicates) ===
            if rag_learnings_created == 0 and original_category != update.category_code:
                try:
                    learning = await rag_integration.learn_from_correction(
                        transaction=txn,
                        original_category=original_category,
                        corrected_category=update.category_code,
                        corrected_by="user",  # Would come from auth in production
                        notes=f"Bulk update of {len(update.transaction_ids)} similar transactions"
                    )
                    if learning:
                        rag_learnings_created += 1
                        logger.info(f"Created RAG learning from bulk correction: {update.category_code}")
                except Exception as e:
                    logger.warning(f"Failed to create RAG learning: {e}")

            updated_count += 1
        except Exception as e:
            logger.error(f"Failed to update transaction {txn_id}: {e}")

    return {
        "updated_count": updated_count,
        "pattern_id": str(pattern_id) if pattern_id else None,
        "apply_to_similar": update.apply_to_similar
    }


@transaction_router.get("/totals/{tax_return_id}")
async def get_transaction_totals(
    tax_return_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """Get summary totals for a tax return (lightweight endpoint for UI updates)."""
    # Get categories to determine income vs expense and deductibility
    cat_result = await db.execute(select(PLRowMapping))
    categories = cat_result.scalars().all()
    # category_map available for future use if needed
    income_categories = {c.category_code for c in categories if c.category_group == 'Income'}
    excluded_categories = {c.category_code for c in categories if c.category_group == 'Excluded'}
    deductible_categories = {c.category_code for c in categories if c.is_deductible}

    result = await db.execute(
        select(Transaction).where(Transaction.tax_return_id == tax_return_id)
    )
    transactions = result.scalars().all()

    needs_review_count = sum(1 for t in transactions if t.needs_review)

    # Total Income (rent, rates recovered, bank contributions, insurance payouts, etc.)
    total_income = sum(t.amount for t in transactions if t.amount > 0 and t.category_code in income_categories)

    # Interest Expense
    interest_expense = sum(abs(t.deductible_amount if t.deductible_amount is not None else t.amount) for t in transactions if t.category_code == 'interest' and t.amount < 0)

    # Other Deductible Expenses - use category's is_deductible flag
    other_deductible = sum(abs(t.deductible_amount if t.deductible_amount is not None else t.amount) for t in transactions if t.amount < 0 and t.category_code in deductible_categories and t.category_code != 'interest')

    # Total Deductible
    total_deductible = float(interest_expense) + float(other_deductible)

    # Net Rental Income
    net_rental_income = float(total_income) - total_deductible

    # Non-Deductible - expenses not in deductible categories or in excluded
    non_deductible = sum(abs(t.amount) for t in transactions if t.amount < 0 and (t.category_code not in deductible_categories or t.category_code in excluded_categories))

    return {
        "needs_review_count": needs_review_count,
        "total_count": len(transactions),
        "total_income": float(total_income),
        "interest_expense": float(interest_expense),
        "other_deductible": float(other_deductible),
        "total_deductible": total_deductible,
        "net_rental_income": net_rental_income,
        "non_deductible": float(non_deductible),
    }


@transaction_router.post("/save-learnings/{tax_return_id}")
async def save_transaction_learnings(
    tax_return_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """
    Save manually reviewed transaction patterns as learnings to RAG (Pinecone).

    Only saves transactions that have been manually reviewed and changed by the user
    (where manually_reviewed=True). Checks for duplicates before saving.
    """
    # Get only MANUALLY REVIEWED transactions (user actually changed/confirmed the category)
    result = await db.execute(
        select(Transaction)
        .where(
            Transaction.tax_return_id == tax_return_id,
            Transaction.manually_reviewed.is_(True)  # Only save user-changed transactions
        )
    )
    transactions = result.scalars().all()

    if not transactions:
        return {
            "saved_count": 0,
            "skipped_count": 0,
            "duplicate_count": 0,
            "message": "No manually reviewed transactions to save. Only transactions you've changed will be saved as learnings."
        }

    # Get category mappings for display names
    cat_result = await db.execute(select(PLRowMapping))
    category_map = {c.category_code: c for c in cat_result.scalars().all()}

    saved_count = 0
    duplicate_count = 0
    skipped_count = 0
    errors = []

    for txn in transactions:
        try:
            # Skip if no category assigned
            if not txn.category_code:
                skipped_count += 1
                continue

            # Get category info
            category = category_map.get(txn.category_code)
            category_name = category.display_name if category else txn.category_code
            is_deductible = category.is_deductible if category else txn.is_deductible

            # Build a unique key for duplicate checking (description + category)
            # Normalize description for comparison
            desc_normalized = (txn.description or "").strip().lower()

            # Check for existing similar learning in Pinecone
            existing = await knowledge_store.search(
                query=f"{desc_normalized} {txn.category_code}",
                top_k=5,
                min_score=0.95,  # High threshold for exact/near-exact matches
                namespace="transaction-coding"
            )

            # Check if we already have a learning for this exact description + category
            is_duplicate = False
            for existing_learning in existing:
                existing_content = existing_learning.get("content", "").lower()
                existing_category = existing_learning.get("category", "")

                # Check if same description pattern and same category
                if desc_normalized[:50] in existing_content and existing_category == txn.category_code:
                    is_duplicate = True
                    logger.debug(f"Duplicate learning found for: {txn.description[:50]}... -> {txn.category_code}")
                    break

            if is_duplicate:
                duplicate_count += 1
                continue

            # Build learning content
            if is_deductible:
                content = (
                    f"Transaction '{txn.description}' (amount: ${abs(txn.amount):.2f}) "
                    f"should be categorized as '{category_name}' (code: {txn.category_code}). "
                    f"This is a DEDUCTIBLE rental property expense. "
                    f"Other party: {txn.other_party or 'N/A'}. "
                    f"Apply similar categorization to transactions with matching descriptions."
                )
                scenario = "deductible_expense_pattern"
            else:
                content = (
                    f"Transaction '{txn.description}' (amount: ${abs(txn.amount):.2f}) "
                    f"should be categorized as '{category_name}' (code: {txn.category_code}). "
                    f"This is NOT deductible for rental property tax purposes. "
                    f"Other party: {txn.other_party or 'N/A'}. "
                    f"Apply similar categorization to transactions with matching descriptions."
                )
                scenario = "non_deductible_pattern"

            # Store in Pinecone
            record_id = await knowledge_store.store(
                content=content,
                scenario=scenario,
                category=txn.category_code or "uncategorized",
                source="transaction_review",
                namespace="transaction-coding"
            )

            if record_id:
                saved_count += 1
                logger.info(f"Saved learning for transaction {txn.id}: {record_id}")

        except Exception as e:
            logger.error(f"Error saving learning for transaction {txn.id}: {e}")
            errors.append(str(txn.id))

    return {
        "saved_count": saved_count,
        "total_reviewed": len(transactions),
        "duplicate_count": duplicate_count,
        "skipped_count": skipped_count,
        "errors": errors if errors else None,
        "message": f"Saved {saved_count} new learnings ({duplicate_count} duplicates skipped)"
    }


@transaction_router.get("/summaries/{tax_return_id}", response_model=List[TransactionSummaryResponse])
async def get_transaction_summaries(
    tax_return_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """Get category summaries for a tax return."""
    result = await db.execute(
        select(TransactionSummary)
        .where(TransactionSummary.tax_return_id == tax_return_id)
        .order_by(TransactionSummary.category_code)
    )
    summaries = result.scalars().all()

    # Get P&L mappings
    mappings_result = await db.execute(select(PLRowMapping))
    mappings = {m.category_code: m for m in mappings_result.scalars().all()}

    return [
        TransactionSummaryResponse(
            id=s.id,
            tax_return_id=s.tax_return_id,
            category_code=s.category_code,
            category_display_name=mappings.get(s.category_code).display_name if mappings.get(s.category_code) else None,
            pl_row=mappings.get(s.category_code).pl_row if mappings.get(s.category_code) else None,
            transaction_type=mappings.get(s.category_code).transaction_type if mappings.get(s.category_code) else None,
            category_group=mappings.get(s.category_code).category_group if mappings.get(s.category_code) else None,
            default_source=mappings.get(s.category_code).default_source if mappings.get(s.category_code) else None,
            transaction_count=s.transaction_count,
            gross_amount=s.gross_amount,
            deductible_amount=s.deductible_amount,
            gst_amount=s.gst_amount,
            monthly_breakdown=s.monthly_breakdown,
            created_at=s.created_at,
            updated_at=s.updated_at
        )
        for s in summaries
    ]


@transaction_router.post("/summaries/{tax_return_id}/regenerate")
async def regenerate_summaries(
    tax_return_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """Regenerate category summaries after corrections."""
    categorizer = get_transaction_categorizer()
    summaries = await categorizer.generate_summaries(db, tax_return_id)

    return {
        "status": "success",
        "summaries_generated": len(summaries)
    }


# =============================================================================
# CATEGORY MANAGEMENT
# =============================================================================

@transaction_router.get("/categories/all", response_model=List[PLRowMappingResponse])
async def get_all_categories(
    db: AsyncSession = Depends(get_db)
):
    """Get all available categories for dropdown."""
    result = await db.execute(
        select(PLRowMapping).order_by(PLRowMapping.sort_order)
    )
    mappings = result.scalars().all()

    return [
        PLRowMappingResponse(
            id=m.id,
            category_code=m.category_code,
            pl_row=m.pl_row,
            display_name=m.display_name,
            category_group=m.category_group,
            transaction_type=m.transaction_type,
            is_deductible=m.is_deductible,
            default_source=m.default_source,
            sort_order=m.sort_order
        )
        for m in mappings
    ]


class CreateCategoryRequest(BaseModel):
    """Request to create a new category."""
    display_name: str
    category_group: str
    is_deductible: bool = True


@transaction_router.post("/categories/create")
async def create_category(
    request: CreateCategoryRequest,
    db: AsyncSession = Depends(get_db)
):
    """Create a new category in the database."""
    import re
    from uuid import uuid4

    # Generate category_code from display_name
    category_code = re.sub(r'[^a-z0-9]+', '_', request.display_name.lower()).strip('_')

    # Check if category already exists
    existing = await db.execute(
        select(PLRowMapping).where(PLRowMapping.category_code == category_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Category '{request.display_name}' already exists")

    # Determine transaction_type based on group
    if request.category_group == "Income":
        transaction_type = "income"
    elif request.category_group == "Excluded":
        transaction_type = "excluded"
    else:
        transaction_type = "expense"

    # Get max sort_order for the group
    max_sort_result = await db.execute(
        select(PLRowMapping.sort_order)
        .where(PLRowMapping.category_group == request.category_group)
        .order_by(PLRowMapping.sort_order.desc())
        .limit(1)
    )
    max_sort = max_sort_result.scalar_one_or_none() or 0

    # Create new category
    new_category = PLRowMapping(
        id=uuid4(),
        category_code=category_code,
        pl_row=None,  # Custom categories don't have P&L row assignments
        display_name=request.display_name,
        category_group=request.category_group,
        transaction_type=transaction_type,
        is_deductible=request.is_deductible,
        default_source="BS",
        sort_order=max_sort + 1
    )

    db.add(new_category)
    await db.commit()
    await db.refresh(new_category)

    logger.info(f"Created new category: {category_code} ({request.display_name}) in group {request.category_group}")

    return {
        "id": str(new_category.id),
        "category_code": new_category.category_code,
        "display_name": new_category.display_name,
        "category_group": new_category.category_group,
        "transaction_type": new_category.transaction_type,
        "is_deductible": new_category.is_deductible,
        "message": "Category created successfully"
    }


# =============================================================================
# PATTERNS MANAGEMENT
# =============================================================================

@transaction_router.get("/patterns", response_model=List[TransactionPatternResponse])
async def get_learned_patterns(
    limit: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db)
):
    """Get learned transaction patterns."""
    result = await db.execute(
        select(TransactionPattern)
        .order_by(TransactionPattern.times_applied.desc())
        .limit(limit)
    )
    patterns = result.scalars().all()

    return [
        TransactionPatternResponse(
            id=p.id,
            description_normalized=p.description_normalized,
            other_party_normalized=p.other_party_normalized,
            category_code=p.category_code,
            confidence=p.confidence,
            times_applied=p.times_applied,
            times_confirmed=p.times_confirmed,
            times_corrected=p.times_corrected,
            is_global=p.is_global,
            client_id=p.client_id,
            source=p.source,
            created_at=p.created_at,
            last_used_at=p.last_used_at
        )
        for p in patterns
    ]


@transaction_router.delete("/patterns/{pattern_id}")
async def delete_pattern(
    pattern_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """Delete a learned pattern."""
    result = await db.execute(
        select(TransactionPattern).where(TransactionPattern.id == pattern_id)
    )
    pattern = result.scalar_one_or_none()

    if not pattern:
        raise HTTPException(status_code=404, detail="Pattern not found")

    await db.delete(pattern)
    await db.commit()

    return {"status": "deleted", "pattern_id": str(pattern_id)}


# =============================================================================
# PROCESSING ENDPOINTS
# =============================================================================

@transaction_router.post("/process/{tax_return_id}", response_model=ProcessingResult)
async def process_all_transactions(
    tax_return_id: UUID,
    use_claude: bool = Query(True),
    db: AsyncSession = Depends(get_db)
):
    """Process all bank statement documents for a tax return."""
    processor = TransactionProcessor()

    result = await processor.process_tax_return_transactions(
        db=db,
        tax_return_id=tax_return_id,
        use_claude=use_claude
    )

    return result


@transaction_router.get("/process/{tax_return_id}/stream")
async def process_transactions_stream(
    tax_return_id: UUID,
    use_claude: bool = Query(True),
):
    """
    Process transactions with real-time progress via Server-Sent Events (SSE).

    Connect to this endpoint to receive progress updates during processing.
    The stream will send JSON events with: stage, progress (0-100), message, detail
    """
    task_id = str(tax_return_id)

    # Check if processing is already in progress for this tax return
    existing_tracker = get_tracker(task_id)
    if existing_tracker and not existing_tracker.is_complete:
        logger.info(f"Processing already in progress for {task_id}, returning existing stream")
        # Return the existing stream instead of starting new processing
        return StreamingResponse(
            existing_tracker.stream(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            }
        )

    # Create new progress tracker
    tracker = create_tracker(task_id)

    async def run_processing():
        """Background task to run processing with progress updates."""
        try:
            async with AsyncSessionLocal() as db:
                processor = TransactionProcessor()
                await processor.process_tax_return_transactions_with_progress(
                    db=db,
                    tax_return_id=tax_return_id,
                    use_claude=use_claude,
                    progress_tracker=tracker
                )
        except Exception as e:
            logger.error(f"Processing error: {e}")
            await tracker.fail(str(e))
        finally:
            # Clean up tracker after a delay (allow client to receive final events)
            await asyncio.sleep(2)
            remove_tracker(task_id)

    # Start processing in background
    asyncio.create_task(run_processing())

    # Return SSE stream
    return StreamingResponse(
        tracker.stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",  # Disable nginx buffering
        }
    )


@transaction_router.post("/process/document", response_model=ProcessingResult)
async def process_document(
    document_id: UUID = Query(...),
    tax_return_id: UUID = Query(...),
    force_reprocess: bool = Query(False),
    db: AsyncSession = Depends(get_db)
):
    """Process a document to extract and categorize transactions."""
    processor = TransactionProcessor()

    result = await processor.process_document(
        db=db,
        document_id=document_id,
        tax_return_id=tax_return_id,
        force_reprocess=force_reprocess
    )

    return result


@transaction_router.post("/process/reprocess", response_model=ProcessingResult)
async def reprocess_transactions(
    tax_return_id: UUID,
    transaction_ids: Optional[List[UUID]] = None,
    db: AsyncSession = Depends(get_db)
):
    """Reprocess transactions with updated rules and patterns."""
    processor = TransactionProcessor()

    result = await processor.reprocess_transactions(
        db=db,
        tax_return_id=tax_return_id,
        transaction_ids=transaction_ids
    )

    return result


@transaction_router.post("/process/learn")
async def learn_from_feedback(
    transaction_id: UUID,
    correct_category: str,
    confidence_adjustment: float = Query(0.1, ge=0.0, le=1.0),
    db: AsyncSession = Depends(get_db)
):
    """Learn from user correction and update patterns."""
    processor = TransactionProcessor()

    success = await processor.learn_from_feedback(
        db=db,
        transaction_id=transaction_id,
        correct_category=correct_category,
        confidence_adjustment=confidence_adjustment
    )

    return {
        "success": success,
        "message": "Pattern learning successful" if success else "Failed to learn pattern"
    }


# =============================================================================
# WORKBOOK GENERATION
# =============================================================================

@transaction_router.post("/workbook/{tax_return_id}")
async def generate_workbook(
    tax_return_id: UUID,
    process_transactions: bool = Query(True, description="Process/categorize transactions before generating workbook"),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate IR3R Excel workbook for a tax return.

    By default, this will first process/categorize any uncategorized transactions
    to ensure the workbook has accurate data.
    """
    from app.services.workbook_generator import get_workbook_generator

    try:
        # Optionally process transactions first to ensure all are categorized
        if process_transactions:
            processor = TransactionProcessor()
            logger.info(f"Processing transactions before workbook generation for tax return {tax_return_id}")
            await processor.process_tax_return_transactions(
                db=db,
                tax_return_id=tax_return_id,
                use_claude=True
            )

        generator = get_workbook_generator()
        filepath = await generator.generate_workbook(db, tax_return_id)

        # Return download info
        filename = filepath.name

        return {
            "status": "success",
            "filename": filename,
            "download_url": f"/api/transactions/workbook/{tax_return_id}/download"
        }

    except Exception as e:
        logger.error(f"Error generating workbook: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@transaction_router.get("/workbook/{tax_return_id}/download")
async def download_workbook(
    tax_return_id: UUID,
    db: AsyncSession = Depends(get_db),
    force_regenerate: bool = False  # Add query param to force regeneration
):
    """Download generated workbook."""
    from app.services.workbook_generator import get_workbook_generator

    generator = get_workbook_generator()

    # Get tax return for filename
    result = await db.execute(
        select(TaxReturn)
        .options(selectinload(TaxReturn.client))
        .where(TaxReturn.id == tax_return_id)
    )
    tax_return = result.scalar_one_or_none()

    if not tax_return:
        raise HTTPException(status_code=404, detail="Tax return not found")

    # Build filename pattern - must match what workbook_generator creates!
    def _sanitize_filename(name: str) -> str:
        return "".join(c if c.isalnum() or c in "_- " else "_" for c in name).replace(" ", "_")

    client_name = _sanitize_filename(tax_return.client.name)
    year = tax_return.tax_year[-2:]  # FY24 -> 24
    filename = f"PTR01_-_Rental_Property_Workbook_-_{client_name}_-_{year}.xlsx"
    filepath = generator.output_dir / filename

    # Log what's happening
    if filepath.exists():
        logger.info(f"Found existing workbook at: {filepath}")
        if force_regenerate:
            logger.info("Force regeneration requested, deleting old file")
            filepath.unlink()  # Delete the old file
        else:
            # Check the sheet structure of existing file
            from openpyxl import load_workbook
            try:
                wb = load_workbook(filepath, read_only=True)
                logger.info(f"Existing workbook sheets: {wb.sheetnames}")
                wb.close()
            except Exception as e:
                logger.error(f"Error reading existing workbook: {e}")

    if not filepath.exists() or force_regenerate:
        logger.info(f"Generating new workbook for tax return {tax_return_id}")
        # Generate new workbook
        filepath = await generator.generate_workbook(db, tax_return_id)

        # Verify the new file
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filepath, read_only=True)
            logger.info(f"Generated workbook sheets: {wb.sheetnames}")
            wb.close()
        except Exception as e:
            logger.error(f"Error reading generated workbook: {e}")

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =============================================================================
# FEEDBACK / CORRECTIONS
# =============================================================================

@transaction_router.post("/feedback", response_model=CategoryFeedbackResponse)
async def submit_feedback(
    feedback: CategoryFeedbackCreate,
    db: AsyncSession = Depends(get_db)
):
    """Submit a category correction."""
    categorizer = get_transaction_categorizer()

    await categorizer.learn_from_correction(
        db=db,
        transaction_id=feedback.transaction_id,
        corrected_category=feedback.corrected_category,
        corrected_by=feedback.corrected_by,
        notes=feedback.notes,
        create_pattern=feedback.create_pattern
    )

    # Get the feedback record
    result = await db.execute(
        select(CategoryFeedback)
        .where(CategoryFeedback.transaction_id == feedback.transaction_id)
        .order_by(CategoryFeedback.corrected_at.desc())
    )
    fb = result.scalars().first()

    return CategoryFeedbackResponse(
        id=fb.id,
        transaction_id=fb.transaction_id,
        original_category=fb.original_category,
        corrected_category=fb.corrected_category,
        corrected_by=fb.corrected_by,
        corrected_at=fb.corrected_at,
        notes=fb.notes,
        pattern_created=fb.pattern_created,
        pattern_id=fb.pattern_id
    )


# =============================================================================
# WEB ROUTES
# =============================================================================

@transaction_web_router.get("/transactions/{tax_return_id}")
async def transactions_page_redirect(
    tax_return_id: UUID,
):
    """
    Redirect to unified workings page.
    The transactions page has been merged into /workings/{id}.
    """
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=f"/workings/{tax_return_id}", status_code=301)